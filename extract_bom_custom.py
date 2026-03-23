"""
제조기록서(BOM) 정보 추출 스크립트 - 시트 인덱스 최종 버전 (수정본)
============================================================
Level 0: 4번(정보), 5번(포장) 시트
Level 1: 5번(포장) 시트
Level 3: 8번, 9번, 10번 시트
Level 2: 11번 ~ 23번 시트
"""

import pandas as pd
import re
import os
from datetime import datetime

# ──────────────── 설정 영역 ────────────────
# 1. 파일 경로 설정
FILE   = r'C:\Users\ENS-1000\Documents\Antigravity\BOM\25BCEPP-001_유효성평가_생산.xlsm'
OUTPUT = r'C:\Users\ENS-1000\Documents\Antigravity\BOM\BOM_25BCEPP-001_유효성평가_생산_최종추출_최종.xlsx'

# 2. 시트 번호(인덱스) 설정 (0부터 시작 기준)
L0_INFO_IDX = 4       # '정보입력' 성격
L1_PACK_IDX = 5       # '포장/완제품' 성격

L3_SHEET_INDICES = [8, 9, 10]          # Level 3 시트 (8, 9, 10)
L2_SHEET_INDICES = list(range(11, 24)) # Level 2 시트 (11 ~ 23번 포함)
# ──────────────────────────────────────────

def get_output_filename(base_path):
    dir_name = os.path.dirname(base_path)
    base_name = os.path.basename(base_path)
    name_no_ext, ext = os.path.splitext(base_name)
    
    i = 1
    while True:
        target = os.path.join(dir_name, f"{name_no_ext}_{i}{ext}")
        if not os.path.exists(target): return target
        i += 1

def to_date_str(val):
    if not val or pd.isna(val): return ''
    if isinstance(val, (pd.Timestamp, datetime)):
        return val.strftime('%Y-%m-%d')
    
    # 엑셀 시리얼 날짜 처리 (숫자형인 경우)
    try:
        num = float(val)
        # 8자리 숫자가 아닌 경우 엑셀 시리얼 날짜로 간주 (예: 44000 ~ 60000 범위)
        if 30000 < num < 60000:
            dt = pd.to_datetime(num, unit='D', origin='1899-12-30')
            return dt.strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        pass

    s = str(val).strip()
    # 숫자만 8자리인 경우 (예: 20240320)
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    
    # 2024.03.20 등 다양한 구분자 처리
    m = re.search(r'(\d{4})[./-]?(\d{1,2})[./-]?(\d{1,2})', s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    
    return s

def cell(df, row, col, is_date=False):
    try:
        if row >= len(df) or col >= len(df.columns): return ''
        v = df.iloc[row, col]
        if pd.isna(v): return ''
        
        if is_date:
            return to_date_str(v)
            
        if isinstance(v, float) and v == int(v): return str(int(v))
        return str(v).strip()
    except Exception:
        return ''

def find_row(df, keyword, col=0):
    for i, v in enumerate(df.iloc[:, col]):
        if pd.notna(v) and keyword in str(v):
            return i
    return -1

def read(sheet_idx):
    return pd.read_excel(FILE, sheet_name=sheet_idx, header=None)

# ──────────────── Level 0 ────────────────
def extract_level0():
    info = read(L0_INFO_IDX)
    fi   = read(L1_PACK_IDX)

    lot_no  = cell(info, 11, 6)
    ref_no  = cell(info, 11, 9)

    prod    = (cell(fi, 6, 0) + ' ' + cell(fi, 6, 4)).strip()
    version = cell(fi, 6, 9)
    pkg_qty = cell(fi, 6, 13)
    mfg     = cell(fi, 6, 18, is_date=True)
    exp     = cell(fi, 6, 25, is_date=True)
    lot_fi  = cell(fi, 6, 30)

    return pd.DataFrame([{
        'Level'          : 0,
        '제품코드'       : ref_no,
        '제품명'         : prod,
        'LOT NO.'        : lot_fi or lot_no,
        '제품버전'       : version,
        '제조일자'       : mfg,
        '유효기간'       : exp,
        '생산 수량(kit)' : pkg_qty,
    }])

# ──────────────── Level 1 ────────────────
def extract_level1(df_inst, master_map):
    df = read(L1_PACK_IDX)
    parent_lot = cell(df, 6, 30)

    # 1. 반제품 맵핑 데이터 확보 (Lot No -> 약어)
    inst_map = {}
    if df_inst is not None and not df_inst.empty:
        for _, row in df_inst.iterrows():
            lot = str(row.get('Lot. No.', '')).strip()
            if lot: inst_map[lot] = row.get('약어', '')

    def get_code_info(name, lot):
        if lot in inst_map: return inst_map[lot], '반제품'
        name_clean = str(name).replace('\n', ' ').strip()
        if master_map and name_clean in master_map: return master_map[name_clean], '원재료'
        if master_map:
            for m_name, m_code in master_map.items():
                if name_clean.startswith(m_name) or m_name.startswith(name_clean): return m_code, '원재료'
        return '', '원재료'

    hdr = find_row(df, '번호', 0)
    if hdr == -1: hdr = 19

    rows = []
    for i in range(hdr + 1, min(hdr + 30, len(df))):
        no_v = cell(df, i, 0)
        name = cell(df, i, 2)
        if not name or not re.match(r'^\d+$', no_v):
            continue

        lot_raw = cell(df, i, 11)
        mfg = cell(df, i, 18, is_date=True)
        
        exp_col = -1
        for c in range(df.shape[1]):
            h = str(df.iloc[hdr, c]) if pd.notna(df.iloc[hdr, c]) else ""
            if "유효기간" in h:
                exp_col = c; break
        
        exp_raw = ""
        search_cols = [exp_col] if exp_col != -1 else [23, 24, 25, 26]
        for c in search_cols:
            v = cell(df, i, c)
            if v and str(v).strip() and not any(kw in str(v) for kw in ["적합", "부적합"]):
                exp_raw = str(v).strip(); break

        lots = re.split(r'[\n\r]+', lot_raw) if lot_raw else ['']
        exps = re.split(r'[\n\r]+', exp_raw) if exp_raw else ['']
        max_lines = max(len(lots), len(exps))

        is_multi = max_lines > 1
        for j in range(max_lines):
            curr_lot = lots[j].strip() if j < len(lots) else (lots[-1].strip() if lots else "")
            curr_exp_str = exps[j].strip() if j < len(exps) else (exps[-1].strip() if exps else "")
            curr_exp = to_date_str(curr_exp_str)

            code_no, gubun = get_code_info(name, curr_lot)

            rows.append({
                'Level'         : 1,
                '상위Lot'       : parent_lot,
                '코드번호'      : code_no,
                '구분'          : gubun,
                '구성품 명칭'   : name,
                'Lot No.'       : curr_lot,
                '제조일자'      : mfg,
                '유효기간'      : curr_exp,
                '포장 기준량'   : cell(df, i, 30),
                '포장시 요구량' : cell(df, i, 34),
                '단위'          : cell(df, i, 37).replace('(', '').replace(')', '').strip(),
                '_multi_lot'    : is_multi
            })
    return pd.DataFrame(rows)

# ──────────────── Level 2/3 (공통 파서) ────────────────
def parse_sheet(sheet_idx, level, df_inst=None, master_map=None):
    try:
        df = read(sheet_idx)
    except Exception: return []

    # 1. 반제품 맵핑 데이터 확보 (Lot No -> 약어)
    inst_map = {}
    if df_inst is not None and not df_inst.empty:
        for _, row in df_inst.iterrows():
            lot = str(row.get('Lot. No.', '')).strip()
            if lot: inst_map[lot] = row.get('약어', '')

    def get_code_info(name, lot):
        if lot in inst_map: return inst_map[lot], '반제품'
        name_clean = str(name).replace('\n', ' ').strip()
        if master_map and name_clean in master_map: return master_map[name_clean], '원재료'
        if master_map:
            for m_name, m_code in master_map.items():
                if name_clean.startswith(m_name) or m_name.startswith(name_clean): return m_code, '원재료'
        return '', '원재료'

    parent_lot, mfg_date = '', ''
    for c in range(df.shape[1]):
        h = str(df.iloc[5, c]) if pd.notna(df.iloc[5, c]) else ''
        if 'Lot No' in h or 'Lot. No' in h: parent_lot = cell(df, 6, c, is_date=False)
        if '제조일자' in h: mfg_date = cell(df, 6, c, is_date=True)

    mat_hdr = find_row(df, '번호', 0)
    if mat_hdr == -1: return []

    hdr_vals = [str(df.iloc[mat_hdr, c]) if pd.notna(df.iloc[mat_hdr, c]) else '' for c in range(df.shape[1])]
    
    def find_col(*kws):
        for kw in kws:
            for ci, h in enumerate(hdr_vals):
                if kw in h: return ci
        return -1

    cc, cm, ct, cl, ce = find_col('Code No', 'Code'), find_col('제조사'), find_col('Cat No'), find_col('Lot No'), find_col('유효기간')

    rows = []
    for i in range(mat_hdr + 1, len(df)):
        no_v = cell(df, i, 0)
        name = cell(df, i, 2)
        if not name: continue
        if no_v and not re.match(r'^[\d\-\.]+$', no_v): break
        if '참고사항' in str(df.iloc[i, 0]) or '계산' in str(df.iloc[i, 0]): break

        def gcol(c): return cell(df, i, c) if c >= 0 else ''

        qty_v, unit_v = '', ''
        for c in range(df.shape[1] - 1, max(cl, 10), -1):
            v = cell(df, i, c)
            if v and v not in ['', '-', 'NaN']:
                if not unit_v: unit_v = v
                elif not qty_v: qty_v = v; break

        # 단위 통일 로직 (ℓ/㎕ -> ㎖)
        u_clean = str(unit_v).replace('(', '').replace(')', '').strip()
        if u_clean in ['ℓ', 'l', 'L', '㎕', 'ul', 'uL']:
            try:
                q_num = float(str(qty_v).replace(',', ''))
                if u_clean in ['ℓ', 'l', 'L']:
                    qty_v = str(q_num * 1000)
                elif u_clean in ['㎕', 'ul', 'uL']:
                    qty_v = str(q_num * 0.001)
                unit_v = '㎖'
            except (ValueError, TypeError):
                pass

        # 생산량(제조량)이 ""(공백)이면 행 삭제
        if not str(qty_v).strip():
            continue

        # 제조량 0인 경우 처리
        try:
            q_num = float(str(qty_v).replace(',', ''))
            if level == 2 and q_num == 0: continue
        except (ValueError, TypeError):
            pass

        lot_raw = gcol(cl)
        exp_raw = cell(df, i, ce, is_date=False) if ce >= 0 else ''
        
        lots = re.split(r'[\n\r]+', lot_raw) if lot_raw else ['']
        exps = re.split(r'[\n\r]+', exp_raw) if exp_raw else ['']
        max_lines = max(len(lots), len(exps))
        is_multi = max_lines > 1
        for j in range(max_lines):
            curr_lot = lots[j].strip() if j < len(lots) else (lots[-1].strip() if lots else "")
            curr_exp_str = exps[j].strip() if j < len(exps) else (exps[-1].strip() if exps else "")
            curr_exp = to_date_str(curr_exp_str)

            code_no, gubun = get_code_info(name, curr_lot)
            
            # Level 2의 경우 원본 엑셀에 코드가 있으면 우선 사용, 없으면 매핑 데이터 사용
            raw_code = gcol(cc)
            final_code = raw_code if raw_code else code_no

            rows.append({
                'Level'   : level,
                '상위Lot' : parent_lot,
                '코드번호': final_code,
                '구분'     : gubun,
                '원재료명': name,
                '제조사'  : gcol(cm),
                'Lot No.' : curr_lot,
                '제조일자': to_date_str(mfg_date),
                '유효기간': curr_exp,
                '제조량'  : qty_v,
                '단위'    : unit_v,
                '_multi_lot': is_multi
            })
    return rows

# ──────────────── 반제품 제조지침서 (요약 시트) ────────────────
def extract_instruction_summary():
    s4 = read(L0_INFO_IDX) # 정보입력 (시트4)
    s5 = read(L1_PACK_IDX) # FI-1(01) (시트5)

    def g(df, row, col, is_date=False):
        return cell(df, row, col, is_date=True) if is_date else cell(df, row, col)

    parent_lot = g(s5, 6, 30) # AE7
    mfg_s5 = g(s5, 6, 18, is_date=True) # S7
    qty_s5 = g(s5, 6, 13) # N7
    inst_s5 = g(s5, 3, 0).replace('제조지침서 :', '').strip() # A4

    rows = []
    def add_item(약어, 지침서, 로트, 수량, 일자):
        if not str(수량).strip(): return
        rows.append({
            '상위Lot': parent_lot, '약어': 약어, '제조지침서 No.': 지침서,
            'Lot. No.': 로트, '생산량': 수량, '제조일자': 일자
        })

    # 1. FI
    add_item('FI', inst_s5, parent_lot, qty_s5, mfg_s5)

    # 2. PL-1, 2, 3
    inst_pl = g(s4, 25, 2)
    for i in range(3):
        add_item(f'PL-{i+1}', inst_pl, g(s4, 27+i, 1), g(s4, 27+i, 3), g(s4, 27+i, 4, is_date=True))

    # 3. SS-1, 2, 3
    inst_ss = g(s4, 25, 8)
    for i in range(3):
        add_item(f'SS-{i+1}', inst_ss, g(s4, 27+i, 6), g(s4, 27+i, 8), g(s4, 27+i, 9, is_date=True))

    # 4. 기타 (CB~TM)
    etc_names = ['CB', 'PB', 'WB', 'CR', 'PC', 'NC', 'DA', 'RD', 'WS', 'TM']
    for i, name in enumerate(etc_names):
        qty = g(s4, 33+i, 4)
        if name in ['CB', 'PB']:
            try: qty = str(float(qty) * 1000)
            except: pass
        add_item(name, g(s4, 33+i, 1), g(s4, 33+i, 2), qty, g(s4, 33+i, 5, is_date=True))

    return pd.DataFrame(rows)

if __name__ == '__main__':
    if not os.path.exists(FILE):
        print(f"❌ 파일을 찾을 수 없습니다: {FILE}")
    else:
        FINAL_OUTPUT = get_output_filename(OUTPUT)
        print(f"🔍 BOM 추출 시작... (저장파일명: {FINAL_OUTPUT})")

        # 0. 마스터 데이터 로드
        master_map = {}
        MASTER_FILE = r'C:\Users\ENS-1000\Documents\Antigravity\BOM\원료마스터.xlsx'
        try:
            mdf = pd.read_excel(MASTER_FILE)
            clean = lambda x: str(x).replace('\n', ' ').strip()
            for _, row in mdf.iterrows():
                code = str(row.get('코드번호', '')).strip()
                name = clean(row.get('제 품 명', ''))
                if name: master_map[name] = code
        except: pass

        df0 = extract_level0()

        # 1. 반제품 제조지침서 추출
        print("  - 반제품 제조지침서 요약 시트 생성 중...")
        df_inst = extract_instruction_summary()

        # 2. Level 1 추출
        print("  - Level 1 보정 추출 중 (코드번호/구분 반영)...")
        df1 = extract_level1(df_inst, master_map)

        # 3. Level 3 추출
        l3_rows = []
        for idx in L3_SHEET_INDICES:
            print(f"  - Level 3 시트 {idx} 처리 중...")
            l3_rows.extend(parse_sheet(idx, 3, df_inst, master_map))
        df3 = pd.DataFrame(l3_rows)

        # 4. Level 2 추출
        l2_rows = []
        for idx in L2_SHEET_INDICES:
            print(f"  - Level 2 시트 {idx} 처리 중...")
            l2_rows.extend(parse_sheet(idx, 2, df_inst, master_map))
        df2 = pd.DataFrame(l2_rows)

        # 5. 엑셀 저장 및 스타일링
        from openpyxl.styles import PatternFill
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        with pd.ExcelWriter(FINAL_OUTPUT, engine='openpyxl') as writer:
            for df_obj, sn in [(df0, 'Level 0'), (df1, 'Level 1'), (df2, 'Level 2'), (df3, 'Level 3'), (df_inst, '반제품 제조지침서')]:
                if '_multi_lot' in df_obj.columns:
                    # 플래그 위치 저장 및 원본 저장
                    is_multi_series = df_obj['_multi_lot'].values
                    df_clean = df_obj.drop(columns=['_multi_lot'])
                    df_clean.to_excel(writer, sheet_name=sn, index=False)
                    
                    ws = writer.sheets[sn]
                    # 행 단위 스타일 적용 (헤더 제외 r_idx=2부터)
                    for r_idx, multi in enumerate(is_multi_series, start=2):
                        if multi:
                            for c_idx in range(1, len(df_clean.columns) + 1):
                                ws.cell(row=r_idx, column=c_idx).fill = yellow_fill
                else:
                    df_obj.to_excel(writer, sheet_name=sn, index=False)

        print(f"💾 저장 완료: {os.path.abspath(FINAL_OUTPUT)}")
