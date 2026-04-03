import calendar
import sqlite3
import pandas as pd
import re
from datetime import date, datetime, timedelta
from flask import Flask, jsonify, send_from_directory, request, send_file
import os
import io
import subprocess
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
import tempfile
import urllib.parse
import zipfile
import xml.etree.ElementTree as ET

app = Flask(__name__, static_folder='.')

# 디렉토리 설정
ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(ROOT_DIR, 'bom.db')

# NAS 제조지침서 루트
# - MI_DOC_BASE: 우선 사용. 여러 후보는 ; 또는 | 로 구분 (예: Z:\...\제조지침서;\\NAS\...\제조지침서)
# - 기본 순서: UNC 먼저(Z:는 서비스/IDE 터미널에서 안 보이는 경우가 많음). MI_DOC_BASE로 순서·경로 조정.
_UNC_MI_BASE = (
    r"\\ens-nas918\회사공유폴더\품질경영시스템 표준문서"
    r"\ESH-7501-생산관리 절차서\지침서\제조지침서"
)
_Z_MI_BASE = (
    r"Z:\회사공유폴더\품질경영시스템 표준문서"
    r"\ESH-7501-생산관리 절차서\지침서\제조지침서"
)

# NAS 표준작업지침서 루트 (모델별 BCE01~ 하위, 각 모델 폴더 안에 R1/R2/… 하위 → 숫자 최대 폴더에 .docx 1부)
# - SWI_DOC_BASE: MI_DOC_BASE 와 같이 ; 로 여러 후보 지정 가능
_UNC_SWI_BASE = (
    r"\\ens-nas918\회사공유폴더\회사공유폴더\품질경영시스템 표준문서"
    r"\표준서\ESH-DHF-BCE-SOP-01 표준작업지침서"
)
_Z_SWI_BASE = (
    r"Z:\회사공유폴더\회사공유폴더\품질경영시스템 표준문서"
    r"\표준서\ESH-DHF-BCE-SOP-01 표준작업지침서"
)

# NAS 품질관리 공정도
# 실제 NAS: 공정도 루트 바로 아래 R0/R1/R2 및 모델명이 파일명에 포함된 xlsx (BCE01 하위 폴더 없음인 경우 많음)
# - QMPC_DOC_BASE: MI/SWI 와 동일하게 ; 로 여러 후보 지정 가능
_UNC_QMPC_BASE = (
    r"\\ens-nas918\회사공유폴더\회사공유폴더\품질경영시스템 표준문서"
    r"\표준서\ESH-PC-BCE-01 품질관리 공정도"
)
_UNC_QMPC_BASE_ALT = (
    r"\\ens-nas918\회사공유폴더\회사공유폴더\품질경영시스템 표준문서"
    r"\표준서\ESH-PC-BCE01 품질관리 공정도"
)
_Z_QMPC_BASE = (
    r"Z:\회사공유폴더\회사공유폴더\품질경영시스템 표준문서"
    r"\표준서\ESH-PC-BCE-01 품질관리 공정도"
)
_Z_QMPC_BASE_ALT = (
    r"Z:\회사공유폴더\회사공유폴더\품질경영시스템 표준문서"
    r"\표준서\ESH-PC-BCE01 품질관리 공정도"
)


def _mi_root_candidates():
    """시도할 제조지침서 상위 폴더 목록 (앞쪽이 우선)."""
    paths = []
    env = (os.environ.get("MI_DOC_BASE") or "").strip()
    if env:
        for part in re.split(r"[;|\n]", env):
            p = part.strip().strip('"').strip("'")
            if p and p not in paths:
                paths.append(p)
    for p in (_UNC_MI_BASE, _Z_MI_BASE):
        if p not in paths:
            paths.append(p)
    return paths


def _swi_root_candidates():
    """표준작업지침서 상위 폴더 후보 (앞쪽이 우선)."""
    paths = []
    env = (os.environ.get("SWI_DOC_BASE") or "").strip()
    if env:
        for part in re.split(r"[;|\n]", env):
            p = part.strip().strip('"').strip("'")
            if p and p not in paths:
                paths.append(p)
    for p in (_UNC_SWI_BASE, _Z_SWI_BASE):
        if p not in paths:
            paths.append(p)
    return paths


def _qmpc_root_candidates():
    """품질관리 공정도 상위 폴더 후보."""
    paths = []
    env = (os.environ.get("QMPC_DOC_BASE") or "").strip()
    if env:
        for part in re.split(r"[;|\n]", env):
            p = part.strip().strip('"').strip("'")
            if p and p not in paths:
                paths.append(p)
    for p in (_UNC_QMPC_BASE, _UNC_QMPC_BASE_ALT, _Z_QMPC_BASE, _Z_QMPC_BASE_ALT):
        if p not in paths:
            paths.append(p)
    return paths


MANUFACTURING_INSTRUCTION_SUBFOLDERS = ("BCE01", "BCE02", "BCE03", "BCE04", "BCEPP")
_MI_FILENAME_REV_RE = re.compile(r"(?i)R(\d+)")
# 표준작업지침서: 모델 폴더 아래 R1, R2, … 하위 폴더 (숫자 최대 = 최신)
_SWI_R_SUBDIR_RE = re.compile(r"(?i)^R(\d+)$")

# BCE 모델 공통: 문서번호 ESH-WS({모델})-7501-NN-Rx · 문서명 (표준 14종)
_MANUFACTURING_INSTRUCTION_CATALOG = (
    ("01", "PBSA Buffer 제조지침서"),
    ("02", "Coating Buffer 제조지침서"),
    ("03", "Washing Buffer(10x) 제조지침서"),
    ("04", "Calibrator 제조지침서"),
    ("05", "Positive Control 제조지침서"),
    ("06", "Negative Control 제조지침서"),
    ("07", "Detection Antibody 제조지침서"),
    ("08", "Antibody Coated 96-well Plate 제조지침서"),
    ("09", "Reagent Dilution Buffer 제조지침서"),
    ("10", "Washing Solution(10x) 제조지침서"),
    ("11", "TMB Solution 제조지침서"),
    ("12", "Stop Solution 제조지침서"),
    ("13", "라벨 출력 및 부착 제조지침서"),
    ("14", "진단키트 포장 제조지침서"),
)


def _max_revision_in_filename(filename: str):
    """파일명에 포함된 R01, R02, R3 등 표기 중 최대 개정 번호."""
    matches = _MI_FILENAME_REV_RE.findall(filename)
    if not matches:
        return None
    return max(int(m) for m in matches)


def _file_matches_instruction_seq(filename: str, seq: str) -> bool:
    """파일명에 7501-{seq} 문서 번호가 포함되는지 (01·12 등 인접 번호 오인 방지)."""
    fn = filename.replace(" ", "")
    if re.search(rf"(?<![0-9])7501-{re.escape(seq)}-R\d+", fn, re.I):
        return True
    if re.search(rf"(?<![0-9])7501-{re.escape(seq)}(?:\.docx|-[^R]|$)", fn, re.I):
        return True
    return False


def _doc_modified_timestamp(doc: dict) -> float:
    try:
        s = doc.get("modified") or ""
        return datetime.fromisoformat(s.replace("Z", "+00:00")).timestamp()
    except Exception:
        return 0.0


_W_MAIN_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
# Issued Date / I+ssued run 분리, 콜론 일반·전각
_ISSUED_DATE_RE = re.compile(
    r"(?:Issued|ssued|Issue)\s*Date\s*[:\uFF1A]\s*(\d{4})[\.\-/](\d{1,2})[\.\-/](\d{1,2})",
    re.I,
)
# 보조: 날짜 직후 Rev (점 유무·개정번호는 별 run)
_ISSUED_BEFORE_REV_RE = re.compile(
    r"(?<![0-9])(\d{4})[\.\-/](\d{1,2})[\.\-/](\d{1,2})\s*Rev\.?\s*\d",
    re.I,
)
# 한글 머리/바닥글 (년·월·일 사이 공백 허용)
_ISSUED_KO_RE = re.compile(
    r"(?:발행일|개정일|제\s*개정일)\s*[:\uFF1A]?\s*"
    r"(\d{4})\s*[\.\-/년]\s*(\d{1,2})\s*[\.\-/월]\s*(\d{1,2})(?:\s*일)?",
    re.I,
)
_PLAIN_DATE_RE = re.compile(r"(?<![0-9])(20\d{2})[\.\-/](\d{1,2})[\.\-/](\d{1,2})(?![0-9])")
# 문서 개정 이력 표의 날짜 셀 (YYYY.MM.DD 등)
_REVISION_TBL_DATE = re.compile(
    r"(?<![0-9])(\d{4})\s*[\.\-/]\s*(\d{1,2})\s*[\.\-/]\s*(\d{1,2})(?![0-9])"
)
# 개정 내용란의 [버전 5.3] · 전각 괄호 허용
_REVISION_BRACKET_VERSION_RE = re.compile(
    r"[\[［]\s*버전\s*([0-9]+(?:\.[0-9]+)*)\s*[\]］]",
    re.I,
)


def _ooxml_local_tag_suffix(tag: str) -> str:
    if not tag:
        return ""
    return tag.rsplit("}", 1)[-1]


def _ooxml_collect_text_doc_order(
    xml_bytes: bytes, max_paragraphs: int | None = None
) -> str:
    """
    w:p 단위로 w:t / instrText / delText 를 모은 뒤 단락 사이에 공백을 넣어 이음.
    (단락 경계 없이 이어붙이면 Issued / Date 가 한 줄로 안 잡히는 경우가 있음)
    네임스페이스 URI가 달라도 로컬 태그명이 t/instrText/delText 이면 수집.
    """
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return ""
    W = _W_MAIN_NS
    p_tag = f"{{{W}}}p"
    paras: list[str] = []
    for p_el in root.iter(p_tag):
        chunks: list[str] = []
        for el in p_el.iter():
            suf = _ooxml_local_tag_suffix(el.tag)
            if suf not in ("t", "instrText", "delText"):
                continue
            if el.text:
                chunks.append(el.text)
        if chunks:
            paras.append("".join(chunks))
        if max_paragraphs is not None and len(paras) >= max_paragraphs:
            break
    joined = " ".join(paras)
    if joined.strip():
        return joined
    # 비정형(단락 밖 run 등): 전체 트리 순회 폴백
    return "".join(
        (el.text or "")
        for el in root.iter()
        if _ooxml_local_tag_suffix(el.tag) in ("t", "instrText", "delText")
    )


def _fmt_yyyy_mm_dd(y: str, mo: str, d: str) -> str:
    return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"


def _yyyy_mm_dd_sort_key(s: str) -> tuple[int, int, int]:
    """YYYY-MM-DD 문자열 비교용 (여러 개정 이력 표 중 최신일 선택)."""
    try:
        parts = s.split("-")
        if len(parts) != 3:
            return (0, 0, 0)
        return (int(parts[0]), int(parts[1]), int(parts[2]))
    except (ValueError, TypeError):
        return (0, 0, 0)


def _revision_history_rows_sort_key(row: dict) -> tuple:
    """개정 이력 행: 제/개정 일자 오름차순, 같은 날은 개정 No. 숫자 오름차순."""
    prim = _yyyy_mm_dd_sort_key(str(row.get("date") or ""))
    rev_raw = str(row.get("revision") or "").strip()
    try:
        rev_n = int(rev_raw)
    except ValueError:
        rev_n = 0
    return (prim, rev_n)


_OOXML_HDR_TAIL_RE = re.compile(
    r"^(?:(?:even|odd|first)header|header)(\d+)\.xml$",
    re.I,
)
_OOXML_FTR_TAIL_RE = re.compile(
    r"^(?:(?:even|odd|first)footer|footer)(\d+)\.xml$",
    re.I,
)


def _issued_part_sort_key(part_name: str) -> tuple[int, int]:
    """
    동일 정규식 우선순위일 때: 머리글 > 바닥글 > 기타 > 본문(document.xml),
    머리글은 번호가 큰 쪽(header4 > header1).
    """
    tail = part_name.rsplit("/", 1)[-1]
    if tail.lower() == "document.xml":
        return (4, 0)
    m = _OOXML_HDR_TAIL_RE.match(tail)
    if m:
        return (0, int(m.group(1)))
    m = _OOXML_FTR_TAIL_RE.match(tail)
    if m:
        return (1, int(m.group(1)))
    if re.match(r"^(endnotes|footnotes)\d*\.xml$", tail, re.I):
        return (2, 0)
    return (3, 0)


def _match_issued_date_with_priority(blob: str):
    """(정규식 우선순위 0=가장 엄격, match) 또는 None."""
    if not blob:
        return None
    b = _normalize_text_for_issued_date(blob)
    for prio, rx in enumerate(
        (_ISSUED_DATE_RE, _ISSUED_KO_RE, _ISSUED_BEFORE_REV_RE)
    ):
        m = rx.search(b)
        if m:
            return (prio, m)
    low = b.lower()
    for key in ("issued", "ssued"):
        pos = low.find(key)
        if pos >= 0:
            window = b[pos : pos + 140]
            m = _PLAIN_DATE_RE.search(window)
            if m:
                return (3, m)
    return None


def _match_issued_date_in_blob(blob: str):
    """blob에서 첫 매칭 (m, groups y,mo,d) 또는 None."""
    r = _match_issued_date_with_priority(blob)
    return r[1] if r else None


def _docx_xml_names_for_issued(zf: zipfile.ZipFile) -> list:
    """머리글·바닥글·각주·미주 (document.xml 제외). even/first 머리글 포함."""
    out = []
    for n in sorted(zf.namelist()):
        if not n.endswith(".xml") or not n.startswith("word/"):
            continue
        tail = n[5:]
        if tail.startswith(("header", "footer", "endnotes", "footnotes")):
            out.append(n)
            continue
        if _OOXML_HDR_TAIL_RE.match(tail) or _OOXML_FTR_TAIL_RE.match(tail):
            out.append(n)
    return out


def _ooxml_collect_w_t_text(xml_bytes: bytes) -> str:
    """호환용: w:t 만 (내부적으로 확장 수집 권장)."""
    return _ooxml_collect_text_doc_order(xml_bytes)


def _normalize_text_for_issued_date(blob: str) -> str:
    """NBSP·전각 공백 등을 일반 공백으로."""
    if not blob:
        return ""
    t = blob.replace("\xa0", " ").replace("\u3000", " ")
    t = t.replace("\uFF1A", ":")
    return t


def _ooxml_direct_children_by_suffix(parent: ET.Element, suffix: str) -> list:
    return [ch for ch in parent if _ooxml_local_tag_suffix(ch.tag) == suffix]


def _ooxml_tc_plain_text(tc: ET.Element) -> str:
    """표 셀 내부 w:t·필드 텍스트를 run 순서대로 이음 (날짜가 run으로 쪼개진 경우 대비)."""
    parts: list[str] = []
    for el in tc.iter():
        if _ooxml_local_tag_suffix(el.tag) in ("t", "instrText", "delText") and el.text:
            parts.append(el.text)
    return "".join(parts)


def _revision_history_date_col_label(compact: str) -> bool:
    """헤더 셀: 제/개정 일자 열인지 (공백 제거·슬래시 정규화 후)."""
    c = re.sub(r"\s+", "", compact)
    c = c.replace("／", "/")
    if "제/개정일자" in c or "제개정일자" in c:
        return True
    return "개정" in c and "일자" in c and ("제" in c or "/" in c)


def _revision_table_last_date(tbl: ET.Element) -> str | None:
    """
    단일 w:tbl에서 '제/개정 일자' 열을 찾고, 개정 No. 열이 숫자인 데이터 행만 스캔해
    마지막 유효 날짜(YYYY-MM-DD)를 반환.
    """
    rows = _ooxml_direct_children_by_suffix(tbl, "tr")
    date_col_idx: int | None = None
    for tr in rows:
        cells_raw = [
            _ooxml_tc_plain_text(tc) for tc in _ooxml_direct_children_by_suffix(tr, "tc")
        ]
        cells = [_normalize_text_for_issued_date(x) for x in cells_raw]
        for ci, ct in enumerate(cells):
            if _revision_history_date_col_label(ct):
                date_col_idx = ci
                break
        if date_col_idx is not None:
            break
    if date_col_idx is None:
        return None
    last_m = None
    for tr in rows:
        cells_raw = [
            _ooxml_tc_plain_text(tc) for tc in _ooxml_direct_children_by_suffix(tr, "tc")
        ]
        cells = [_normalize_text_for_issued_date(x) for x in cells_raw]
        if len(cells) <= date_col_idx:
            continue
        r0 = cells[0].strip()
        if not re.match(r"^\d{1,4}$", r0):
            continue
        date_cell = cells[date_col_idx]
        m = _REVISION_TBL_DATE.search(date_cell)
        if m:
            last_m = m
    if last_m:
        return _fmt_yyyy_mm_dd(last_m.group(1), last_m.group(2), last_m.group(3))
    return None


def _revision_table_data_rows(tbl: ET.Element) -> list[dict] | None:
    """
    개정 이력 표 1개에서 데이터 행만 추출.
    반환 None: 이 표가 개정 이력 형식이 아님. []: 형식은 맞으나 데이터 행 없음.
    """
    rows_el = _ooxml_direct_children_by_suffix(tbl, "tr")
    date_col_idx: int | None = None
    for tr in rows_el:
        cells_raw = [
            _ooxml_tc_plain_text(tc) for tc in _ooxml_direct_children_by_suffix(tr, "tc")
        ]
        cells = [_normalize_text_for_issued_date(x) for x in cells_raw]
        for ci, ct in enumerate(cells):
            if _revision_history_date_col_label(ct):
                date_col_idx = ci
                break
        if date_col_idx is not None:
            break
    if date_col_idx is None:
        return None
    out: list[dict] = []
    for tr in rows_el:
        cells_raw = [
            _ooxml_tc_plain_text(tc) for tc in _ooxml_direct_children_by_suffix(tr, "tc")
        ]
        cells = [_normalize_text_for_issued_date(x) for x in cells_raw]
        if len(cells) <= date_col_idx:
            continue
        r0 = cells[0].strip()
        if not re.match(r"^\d{1,4}$", r0):
            continue
        date_cell = cells[date_col_idx]
        m = _REVISION_TBL_DATE.search(date_cell)
        if not m:
            continue
        date_fmt = _fmt_yyyy_mm_dd(m.group(1), m.group(2), m.group(3))
        rest = cells[date_col_idx + 1 :]
        content = ""
        approval = ""
        if len(rest) >= 4:
            content = rest[0].strip()
            approval = " ".join(x.strip() for x in rest[1:4] if x.strip())
        elif len(rest) >= 1:
            content = rest[0].strip()
            approval = " ".join(x.strip() for x in rest[1:] if x.strip())
        out.append(
            {
                "revision": r0,
                "date": date_fmt,
                "content": content,
                "approval": approval,
            }
        )
    return out


def _revision_history_rows_from_document_xml(xml_bytes: bytes) -> list[dict]:
    """document.xml에서 개정 이력 표(들)의 데이터 행을 본문 순서대로 이어 붙임."""
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return []
    body = None
    for el in root.iter():
        if _ooxml_local_tag_suffix(el.tag) == "body":
            body = el
            break
    if body is None:
        return []
    combined: list[dict] = []
    for tbl in body.iter():
        if _ooxml_local_tag_suffix(tbl.tag) != "tbl":
            continue
        chunk = _revision_table_data_rows(tbl)
        if chunk is not None:
            combined.extend(chunk)
    combined.sort(key=_revision_history_rows_sort_key)
    return combined


def _bracket_doc_version_from_revision_rows(rows: list[dict]) -> str | None:
    """
    일자 오름차순 정렬된 개정 이력 행에서, 가장 최신(마지막) 쪽부터 [버전 x.y] 검색.
    """
    if not rows:
        return None
    for r in reversed(rows):
        m = _REVISION_BRACKET_VERSION_RE.search((r.get("content") or ""))
        if m:
            return m.group(1).strip()
    return None


def _bracket_doc_version_from_revision_docx_path(path: str) -> str | None:
    """docx 본문 개정 이력 표에서 문서 버전 문자열(예: 5.3)만 추출."""
    if not path or not isinstance(path, str):
        return None
    try:
        with open(path, "rb") as f:
            raw = f.read()
    except OSError:
        return None
    if not raw or len(raw) < 4 or raw[:2] != b"PK":
        return None
    try:
        with zipfile.ZipFile(io.BytesIO(raw), "r") as zf:
            if "word/document.xml" not in zf.namelist():
                return None
            xml_bytes = zf.read("word/document.xml")
    except (zipfile.BadZipFile, OSError, KeyError, RuntimeError):
        return None
    rows = _revision_history_rows_from_document_xml(xml_bytes)
    return _bracket_doc_version_from_revision_rows(rows)


def _safe_instruction_docx_path(folder_key: str, filename: str, roots: list) -> str | None:
    """NAS 지정 루트 하위 폴더 안의 .docx 절대경로만 허용 (경로 탈출 방지)."""
    if not filename or not folder_key:
        return None
    fn = filename.strip()
    if not fn or os.path.basename(fn) != fn or "/" in fn or "\\" in fn:
        return None
    base = fn
    if not base.lower().endswith(".docx") or base.startswith("~$"):
        return None
    folder_path, _, _ = _resolve_instruction_subfolder(folder_key, roots)
    if not folder_path:
        return None
    full = os.path.normpath(os.path.join(folder_path, base))
    root_norm = os.path.normpath(folder_path)
    try:
        if os.path.commonpath([full, root_norm]) != root_norm:
            return None
    except ValueError:
        return None
    if not os.path.isfile(full):
        return None
    return full


def _safe_mi_docx_path(folder_key: str, filename: str) -> str | None:
    return _safe_instruction_docx_path(folder_key, filename, _mi_root_candidates())


def _safe_swi_docx_path(folder_key: str, filename: str) -> str | None:
    """
    SWI: 모델 폴더 바로 아래(구형) 또는 숫자가 가장 큰 R* 하위 폴더 안의 .docx만 허용.
    파일명이 여러 R 폴더에 있으면 R번호가 큰 쪽을 우선합니다.
    """
    if not filename or not folder_key:
        return None
    fn = filename.strip()
    if not fn or os.path.basename(fn) != fn or "/" in fn or "\\" in fn:
        return None
    if not fn.lower().endswith(".docx") or fn.startswith("~$"):
        return None
    folder_path, _, _ = _resolve_instruction_subfolder(folder_key, _swi_root_candidates())
    if not folder_path:
        return None

    def _is_file_under(child: str, parent: str) -> bool:
        child_n = os.path.normpath(child)
        parent_n = os.path.normpath(parent)
        try:
            return os.path.commonpath([child_n, parent_n]) == parent_n
        except ValueError:
            return False

    def try_join(root: str) -> str | None:
        full = os.path.normpath(os.path.join(root, fn))
        if _is_file_under(full, root) and os.path.isfile(full):
            return full
        return None

    p = try_join(folder_path)
    if p:
        return p
    r_path, _ = _swi_latest_r_folder_path(folder_path)
    if r_path:
        p = try_join(r_path)
        if p:
            return p
    work, names = _folder_work_path_and_names(folder_path)
    if work and names:
        r_dirs: list[tuple[int, str]] = []
        for name in names:
            m = _SWI_R_SUBDIR_RE.match(name.strip())
            if not m:
                continue
            cand = os.path.join(work, name)
            try:
                if os.path.isdir(cand):
                    r_dirs.append((int(m.group(1)), cand))
            except OSError:
                continue
        r_dirs.sort(key=lambda x: -x[0])
        for _rev, rp in r_dirs:
            p = try_join(rp)
            if p:
                return p
    return None


def _safe_qmpc_xlsx_path(folder_key: str, filename: str) -> str | None:
    """품질관리 공정도: 레거시 BCE01/R* 경로 또는 공정도 루트·R* 평면 구조에서 파일 검증."""
    if not filename or not folder_key:
        return None
    fn = filename.strip()
    if not fn or os.path.basename(fn) != fn or "/" in fn or "\\" in fn:
        return None
    low = fn.lower()
    if not (low.endswith(".xlsx") or low.endswith(".xlsm")) or fn.startswith("~$"):
        return None
    roots = _qmpc_root_candidates()

    def _is_file_under(child: str, parent: str) -> bool:
        child_n = os.path.normpath(child)
        parent_n = os.path.normpath(parent)
        try:
            return os.path.commonpath([child_n, parent_n]) == parent_n
        except ValueError:
            return False

    def try_under(parent: str, scope: str, require_model_match: bool) -> str | None:
        full = os.path.normpath(os.path.join(parent, fn))
        if not _is_file_under(full, scope) or not os.path.isfile(full):
            return None
        if require_model_match and not _qmpc_filename_matches_model(fn, folder_key):
            return None
        return full

    folder_path, _, _ = _resolve_instruction_subfolder(folder_key, roots)
    if folder_path:
        p = try_under(folder_path, folder_path, False)
        if p:
            return p
        r_path, _ = _swi_latest_r_folder_path(folder_path)
        if r_path:
            p = try_under(r_path, folder_path, False)
            if p:
                return p
        work, names = _folder_work_path_and_names(folder_path)
        if work and names:
            r_dirs: list[tuple[int, str]] = []
            for name in names:
                m = _SWI_R_SUBDIR_RE.match(name.strip())
                if not m:
                    continue
                cand = os.path.join(work, name)
                try:
                    if os.path.isdir(cand):
                        r_dirs.append((int(m.group(1)), cand))
                except OSError:
                    continue
            r_dirs.sort(key=lambda x: -x[0])
            for _rev, rp in r_dirs:
                p = try_under(rp, folder_path, False)
                if p:
                    return p

    for root in roots:
        base, _ = _qmpc_accessible_base(root)
        if not base:
            continue
        r_path, _ = _swi_latest_r_folder_path(base)
        for parent in [x for x in (r_path, base) if x]:
            p = try_under(parent, base, True)
            if p:
                return p
    return None


def _issued_date_from_revision_table(xml_bytes: bytes) -> str | None:
    """
    document.xml에서「제/개정 일자」열이 있는 개정 이력 표를 모두 찾아,
    각 표의 마지막 유효 날짜 후보를 모은 뒤 달력상 가장 늦은 날짜를 반환.
    BCE01·BCE03·BCEPP 등 다음 페이지에 이어지는 표가 최신인 서식에 대응.
    """
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return None
    body = None
    for el in root.iter():
        if _ooxml_local_tag_suffix(el.tag) == "body":
            body = el
            break
    if body is None:
        return None
    candidates: list[str] = []
    for tbl in body.iter():
        if _ooxml_local_tag_suffix(tbl.tag) != "tbl":
            continue
        d = _revision_table_last_date(tbl)
        if d:
            candidates.append(d)
    if not candidates:
        return None
    return max(candidates, key=_yyyy_mm_dd_sort_key)


def _issued_date_from_docx_headers(path: str) -> str | None:
    """
    제/개정 일자: (1) 본문에 개정 이력 표가 여러 개면 각 표의 마지막 유효 일자 중 가장 늦은 날,
    (2) 없으면 머리글·바닥글·각주·미주 및 본문 앞단에서 Issued Date 등 휴리스틱.
    """
    if not path or not isinstance(path, str):
        return None
    raw_doc: bytes | None = None
    try:
        with open(path, "rb") as f:
            raw_doc = f.read()
    except OSError:
        return None
    if not raw_doc or len(raw_doc) < 4 or raw_doc[:2] != b"PK":
        return None
    try:
        with zipfile.ZipFile(io.BytesIO(raw_doc), "r") as zf:
            doc_part = (
                "word/document.xml" if "word/document.xml" in zf.namelist() else None
            )
            doc_xml_bytes: bytes | None = None
            if doc_part:
                try:
                    doc_xml_bytes = zf.read(doc_part)
                    t_rev = _issued_date_from_revision_table(doc_xml_bytes)
                    if t_rev:
                        return t_rev
                except KeyError:
                    doc_xml_bytes = None
            part_names_hf = _docx_xml_names_for_issued(zf)
            blobs: dict[str, str] = {}
            for part in part_names_hf:
                try:
                    raw = zf.read(part)
                except KeyError:
                    continue
                blobs[part] = _ooxml_collect_text_doc_order(raw)
            big_hf = _normalize_text_for_issued_date(
                "".join(blobs[p] for p in sorted(blobs))
            )
            blob_doc = ""
            if doc_xml_bytes is not None:
                blob_doc = _ooxml_collect_text_doc_order(
                    doc_xml_bytes, max_paragraphs=120
                )

            candidates: list[tuple[tuple[int, int, int], re.Match]] = []
            for part, blob in blobs.items():
                r = _match_issued_date_with_priority(blob)
                if r:
                    prio, m = r
                    kind, idx = _issued_part_sort_key(part)
                    candidates.append(((prio, kind, -idx), m))

            if doc_part and blob_doc.strip():
                r = _match_issued_date_with_priority(blob_doc)
                if r:
                    prio, m = r
                    kind, idx = _issued_part_sort_key(doc_part)
                    candidates.append(((prio, kind, -idx), m))

            r = _match_issued_date_with_priority(big_hf)
            if r:
                prio, m = r
                candidates.append(((prio, 0, 0), m))

            big_all = _normalize_text_for_issued_date(
                (big_hf + " " + blob_doc) if blob_doc else big_hf
            )
            for m2 in _PLAIN_DATE_RE.finditer(big_hf):
                tail = big_hf[m2.end() : m2.end() + 48]
                if re.search(r"Rev\.?\s*\d", tail, re.I):
                    candidates.append(((4, 0, 0), m2))
                    break
            if blob_doc.strip() and big_all != big_hf:
                for m2 in _PLAIN_DATE_RE.finditer(big_all):
                    tail = big_all[m2.end() : m2.end() + 48]
                    if re.search(r"Rev\.?\s*\d", tail, re.I):
                        candidates.append(((4, 1, 0), m2))
                        break

            if not candidates:
                r = _match_issued_date_with_priority(big_all)
                if r:
                    prio, m = r
                    candidates.append(((prio, 6, 0), m))

            if candidates:
                _, best = min(candidates, key=lambda x: x[0])
                y, mo, d = best.group(1), best.group(2), best.group(3)
                return _fmt_yyyy_mm_dd(y, mo, d)
    except (zipfile.BadZipFile, OSError, RuntimeError):
        return None
    return None


def _build_catalog_rows(model_code: str, all_docs: list, catalog: tuple | None = None) -> list:
    """
    스프레드시트 형식: 모델명, 문서번호, 문서명, 제/개정일, 버전.
    NAS 파일명에서 7501-NN 및 R개정을 찾아 버전·문서번호를 채움.
    제/개정일은 매칭된 docx에서 개정 이력 표 우선, 없으면 머리글·본문 등에서 읽음.
    catalog: 기본 제조지침서 14종 (표준작업지침서는 별도 SWI 페이로드).
    """
    rows_out = []
    cat = catalog if catalog is not None else _MANUFACTURING_INSTRUCTION_CATALOG
    for seq, title in cat:
        matches = [d for d in all_docs if _file_matches_instruction_seq(d["filename"], seq)]
        best = None
        for d in matches:
            r = d.get("revision")
            if r is None:
                continue
            ts = _doc_modified_timestamp(d)
            if best is None or r > best[0] or (r == best[0] and ts > best[1]):
                best = (r, ts, d)
        if best is not None:
            rev, _ts, d = best
            fp = d.get("full_path") or ""
            idate = _issued_date_from_docx_headers(fp)
            rows_out.append(
                {
                    "model_name": model_code,
                    "document_number": f"ESH-WS({model_code})-7501-{seq}-R{rev}",
                    "document_title": title,
                    "issue_revision_date": idate,
                    "version": f"R{rev}",
                    "matched_filename": d["filename"],
                }
            )
        elif matches:
            d = max(matches, key=_doc_modified_timestamp)
            fp = d.get("full_path") or ""
            idate = _issued_date_from_docx_headers(fp)
            rows_out.append(
                {
                    "model_name": model_code,
                    "document_number": f"ESH-WS({model_code})-7501-{seq}",
                    "document_title": title,
                    "issue_revision_date": idate,
                    "version": "—",
                    "matched_filename": d["filename"],
                }
            )
        else:
            rows_out.append(
                {
                    "model_name": model_code,
                    "document_number": f"ESH-WS({model_code})-7501-{seq}",
                    "document_title": title,
                    "issue_revision_date": None,
                    "version": "—",
                    "matched_filename": None,
                }
            )
    return rows_out


def _folder_work_path_and_names(folder_path: str):
    """폴더 내 파일·폴더 이름 목록. (작업경로, 이름리스트) 또는 (None, None)."""
    work_path = folder_path
    names = None
    for candidate in _win_path_access_variants(folder_path):
        try:
            if os.path.isdir(candidate):
                names = os.listdir(candidate)
                work_path = candidate
                break
        except OSError:
            continue
    if names is None and os.name == "nt":
        names, _cmd_err = _win_dir_list_names_cmd(folder_path)
        if names is None:
            for candidate in list(_win_path_access_variants(folder_path))[1:]:
                names, _cmd_err = _win_dir_list_names_cmd(candidate)
                if names is not None:
                    work_path = candidate
                    break
        else:
            work_path = folder_path
    if names is None:
        return None, None
    return work_path, names


def _swi_latest_r_folder_path(model_folder_path: str) -> tuple[str | None, int | None]:
    """모델 폴더 직계 자식 중 이름이 R+숫자인 폴더만 보고, 숫자가 가장 큰 경로를 반환."""
    work, names = _folder_work_path_and_names(model_folder_path)
    if not work or not names:
        return None, None
    best_path: str | None = None
    best_n = -1
    for name in names:
        m = _SWI_R_SUBDIR_RE.match(name.strip())
        if not m:
            continue
        n = int(m.group(1))
        candidate = os.path.join(work, name)
        try:
            if not os.path.isdir(candidate):
                continue
        except OSError:
            continue
        if n > best_n:
            best_n = n
            best_path = candidate
    if best_path is None:
        return None, None
    return best_path, best_n


def _scan_all_docx_in_folder(folder_path: str):
    """
    폴더 내 모든 .docx 목록. R개정 없는 파일도 포함.
    반환: (문서 dict 목록, None, None) 또는 (None, err_code, err_msg)
    """
    work_path, names = _folder_work_path_and_names(folder_path)
    if names is None:
        return None, "not_found", "폴더에 접근할 수 없거나 존재하지 않습니다."

    rows = []
    for name in names:
        if not name.lower().endswith(".docx"):
            continue
        if name.startswith("~$"):
            continue
        full = os.path.join(work_path, name)
        try:
            st = os.stat(full)
            mtime = st.st_mtime
        except OSError:
            continue
        rev = _max_revision_in_filename(name)
        rows.append(
            {
                "filename": name,
                "full_path": full,
                "revision": rev,
                "modified": datetime.fromtimestamp(mtime).isoformat(timespec="seconds"),
            }
        )

    def _sort_key(r):
        rv = r["revision"]
        return (rv is None, -(rv if rv is not None else 0), r["filename"].lower())

    rows.sort(key=_sort_key)
    return rows, None, None


def _scan_all_xlsx_in_folder(folder_path: str):
    """폴더 내 .xlsx / .xlsm 목록 (~$ 잠금 파일 제외)."""
    work_path, names = _folder_work_path_and_names(folder_path)
    if names is None:
        return None, "not_found", "폴더에 접근할 수 없거나 존재하지 않습니다."

    rows = []
    for name in names:
        low = name.lower()
        if not (low.endswith(".xlsx") or low.endswith(".xlsm")):
            continue
        if name.startswith("~$"):
            continue
        full = os.path.join(work_path, name)
        try:
            st = os.stat(full)
            mtime = st.st_mtime
        except OSError:
            continue
        rev = _max_revision_in_filename(name)
        rows.append(
            {
                "filename": name,
                "full_path": full,
                "revision": rev,
                "modified": datetime.fromtimestamp(mtime).isoformat(timespec="seconds"),
            }
        )

    def _sort_key_x(r):
        rv = r["revision"]
        return (rv is None, -(rv if rv is not None else 0), r["filename"].lower())

    rows.sort(key=_sort_key_x)
    return rows, None, None


def _xlsx_cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0:
            return v.strftime("%Y-%m-%d")
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return v.isoformat()
    t = str(v).strip()
    return t


def _qmpc_sheet_has_revision_history(ws) -> bool:
    """시트 상단~중단에 '문서 개정 이력' 또는 개정/일자 열이 있는지."""
    blob = ""
    n = 0
    for row in ws.iter_rows(values_only=True, max_row=85):
        parts = [_xlsx_cell_str(c) for c in row]
        if not any(p.strip() for p in parts):
            continue
        blob += " ".join(parts) + " "
        n += 1
        if n > 60:
            break
    b = blob.replace(" ", "")
    if "문서개정이력" in b:
        return True
    if "개정" in blob and ("일자" in blob or "제/개정" in blob):
        return True
    return False


def _qmpc_pick_revision_sheet_index(wb) -> int | None:
    """Sheet1(첫 탭) 우선 — 개정 이력 표가 있는 첫 시트(최대 앞쪽 8개 탭 검사)."""
    names = list(wb.sheetnames)
    for idx in range(min(8, len(names))):
        try:
            if _qmpc_sheet_has_revision_history(wb[names[idx]]):
                return idx
        except Exception:
            continue
    return None


def _row_looks_like_qmpc_revision_header(cells: list[str]) -> bool:
    """표 제목 행(문서 개정 이력만 있는 행)이 아닌, 열 헤더 행."""
    if not cells or not any(c.strip() for c in cells):
        return False
    joined_ns = "".join(cells).replace(" ", "").replace("\n", "")
    if "문서개정이력" in joined_ns and len([c for c in cells if c.strip()]) <= 2:
        return False
    has_rev = any(
        "개정" in c and ("no" in c.lower() or "번호" in c or "No." in c)
        for c in cells
    )
    has_date = any("일자" in c or "제/개정" in c for c in cells)
    return has_rev and has_date


def _qmpc_raw_nonempty_rows_from_ws(ws) -> list[list[str]]:
    raw_rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = [_xlsx_cell_str(c) for c in row]
        if not any(x.strip() for x in cells):
            continue
        raw_rows.append(cells)
    return raw_rows


def _qmpc_headers_and_body_from_raw(raw_rows: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    """표 위에 표지·제목 행이 있어도 '개정 No / 제·개정 일자' 헤더 행을 찾음."""
    if not raw_rows:
        return [], []
    hi = None
    for i, row in enumerate(raw_rows):
        if _row_looks_like_qmpc_revision_header(row):
            hi = i
            break
    if hi is None:
        hi = 0
    header_row = raw_rows[hi]
    body_slice = raw_rows[hi + 1 :]
    max_len = max(
        [len(header_row)] + [len(r) for r in body_slice],
        default=len(header_row),
    )
    headers = header_row + [""] * (max_len - len(header_row))
    body = [r + [""] * (max_len - len(r)) for r in body_slice]
    return headers, body


def _xlsx_qmpc_revision_history_grid(path: str) -> dict:
    """
    개정 이력이 있는 시트(통상 Sheet1, 일부 서식은 2번째 탭)를 그리드로 읽음.
    """
    out: dict = {"headers": [], "rows": [], "sheet_name": "", "error": None}
    if not path or not os.path.isfile(path):
        out["error"] = "not_found"
        return out
    wb = None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        out["error"] = f"load_error:{e}"
        return out
    try:
        names = list(wb.sheetnames)
        idx = _qmpc_pick_revision_sheet_index(wb)
        if idx is None:
            out["error"] = "revision_sheet_not_found"
            return out
        sn = names[idx]
        out["sheet_name"] = sn
        ws = wb[sn]
        raw_rows = _qmpc_raw_nonempty_rows_from_ws(ws)
        if not raw_rows:
            return out
        headers, body = _qmpc_headers_and_body_from_raw(raw_rows)
        out["headers"] = headers
        out["rows"] = body
        return out
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


# 하위 호환: 기존 이름 유지
def _xlsx_second_sheet_revision_grid(path: str) -> dict:
    return _xlsx_qmpc_revision_history_grid(path)


# 품질관리 공정도 표지/본문 셀에서 추출 (예: ESH-PC-BCE01-01-R2, BCEPP)
_QMPC_DOCNO_RE = re.compile(
    r"ESH-PC-BCE(?:\d{2}|PP)(?:-[A-Za-z0-9]+)*",
    re.IGNORECASE,
)


def _flexible_qmpc_date_to_iso(v) -> str | None:
    """엑셀 개정 이력 일자 → YYYY-MM-DD."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    m = re.match(
        r"(\d{4})\s*[\.\-\/년]?\s*(\d{1,2})\s*[\.\-\/월]?\s*(\d{1,2})",
        s,
    )
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d).isoformat()
        except ValueError:
            return None
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return None


def _qmpc_parse_revision_cell(v) -> int | None:
    t = _xlsx_cell_str(v).strip()
    if not t:
        return None
    m = re.match(r"R\s*(\d+)\s*$", t, re.IGNORECASE)
    if m:
        return int(m.group(1))
    if t.isdigit():
        return int(t)
    m2 = re.search(r"(\d+)", t)
    if m2:
        return int(m2.group(1))
    return None


def _qmpc_revision_latest_from_grid(headers: list, rows: list) -> tuple[int | None, str | None]:
    """개정 이력 표에서 최대 개정 번호와 그 행의 제/개정 일자."""
    if not headers or not rows:
        return None, None
    h = [(x or "").replace("\n", " ").strip() for x in headers]
    rev_i = None
    date_i = None
    for i, cell in enumerate(h):
        c = (cell or "").replace(" ", "")
        if "개정" in c and ("no" in c.lower() or "번호" in c):
            rev_i = i
        if "일자" in (cell or "") or "제/개정" in (cell or ""):
            date_i = i
    if rev_i is None:
        rev_i = 0
    if date_i is None:
        for i, cell in enumerate(h):
            if "일자" in (cell or ""):
                date_i = i
                break
        if date_i is None and len(h) > 1:
            date_i = 1
    best_rev = -1
    best_date: str | None = None
    for row in rows:
        if not row or rev_i >= len(row):
            continue
        rv = _qmpc_parse_revision_cell(row[rev_i])
        if rv is None:
            continue
        ds = row[date_i] if date_i is not None and date_i < len(row) else ""
        dt = _flexible_qmpc_date_to_iso(ds)
        if rv > best_rev:
            best_rev = rv
            best_date = dt
        elif rv == best_rev and dt is not None:
            best_date = dt
    if best_rev < 0:
        return None, None
    return best_rev, best_date


def _qmpc_meta_from_xlsx(path: str) -> dict:
    """
    공정도 xlsx: 표지 등에서 문서번호(ESH-PC-…), 개정 이력 시트(Sheet1 우선)에서 최신 제/개정일·개정번호.
    제/개정일은 NAS 파일 수정일과 무관하게 표 데이터만 사용.
    """
    out: dict = {
        "document_number": None,
        "latest_revision_date": None,
        "latest_revision_no": None,
        "sheet_revision_name": None,
    }
    if not path or not os.path.isfile(path):
        return out
    wb = None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return out
    try:
        names = list(wb.sheetnames)
        candidates: list[str] = []
        for idx in (0, 1, 2):
            if idx >= len(names):
                break
            ws = wb[names[idx]]
            for row in ws.iter_rows(values_only=True, max_row=45):
                for c in row:
                    t = _xlsx_cell_str(c)
                    for m in _QMPC_DOCNO_RE.finditer(t):
                        candidates.append(m.group(0))
        if candidates:
            out["document_number"] = max(candidates, key=len)
        idx = _qmpc_pick_revision_sheet_index(wb)
        if idx is None:
            return out
        sn = names[idx]
        out["sheet_revision_name"] = sn
        ws = wb[sn]
        raw_rows = _qmpc_raw_nonempty_rows_from_ws(ws)
        if not raw_rows:
            return out
        hdr, body = _qmpc_headers_and_body_from_raw(raw_rows)
        rev_no, rev_date = _qmpc_revision_latest_from_grid(hdr, body)
        out["latest_revision_no"] = rev_no
        out["latest_revision_date"] = rev_date
        return out
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _scan_latest_docx_by_revision(folder_path: str):
    """
    폴더 내 .docx 중 파일명의 R개정 번호가 가장 큰 파일을 선택.
    동일 개정이 여러 개면 수정 시각(mtime)이 더 최근인 것을 선택.
    """
    work_path, names = _folder_work_path_and_names(folder_path)
    if names is None:
        return None, "not_found", "폴더에 접근할 수 없거나 존재하지 않습니다."

    best = None  # (revision, mtime, filename, full_path)
    for name in names:
        if not name.lower().endswith(".docx"):
            continue
        if name.startswith("~$"):
            continue
        rev = _max_revision_in_filename(name)
        if rev is None:
            continue
        full = os.path.join(work_path, name)
        try:
            st = os.stat(full)
            mtime = st.st_mtime
        except OSError:
            continue
        if best is None or rev > best[0] or (rev == best[0] and mtime > best[1]):
            best = (rev, mtime, name, full)

    if best is None:
        return None, "no_matching_docx", "R개정 표기가 있는 .docx 파일이 없습니다."
    return best, None, None


def _win_path_access_variants(path: str):
    """Windows에서 동일 경로에 대해 일반 경로와 \\\\?\\ 확장 경로를 순서대로 반환."""
    path = os.path.normpath(path)
    yield path
    if os.name != "nt" or path.startswith("\\\\?\\"):
        return
    if path.startswith("\\\\"):
        # UNC: \\server\share\... → \\?\UNC\server\share\...
        yield "\\\\?\\UNC\\" + path[2:]
    else:
        # 드라이브: Z:\... → \\?\Z:\...
        yield "\\\\?\\" + path


def _win_dir_list_names_cmd(folder_path: str):
    """
    Windows: 탐색기에서는 열리지만 Python os.listdir만 실패하는 환경 대비.
    cmd /c dir /b 로 이름만 나열 (같은 사용자 네트워크 연결을 타는 경우가 있음).
    성공 시 (이름 목록, None), 실패 시 (None, 오류 메시지).
    """
    if os.name != "nt":
        return None, "Windows 전용"
    creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    try:
        si = subprocess.STARTUPINFO()
        si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    except (AttributeError, TypeError):
        si = None
    try:
        r = subprocess.run(
            ["cmd", "/c", "dir", "/b", folder_path],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
            startupinfo=si,
            creationflags=creationflags,
        )
    except subprocess.TimeoutExpired:
        return None, "dir 명령 시간 초과"
    except OSError as e:
        return None, str(e)
    if r.returncode != 0:
        msg = (r.stderr or r.stdout or "").strip() or f"dir 종료 코드 {r.returncode}"
        return None, msg
    names = [ln.strip() for ln in r.stdout.splitlines() if ln.strip()]
    return names, None


def _win_dir_list_subdirs_cmd(parent_path: str):
    """Windows: 하위 폴더 이름만 (dir /ad /b). NAS에 BCE01(96-well) 형태로 올라간 경우 구분용."""
    if os.name != "nt":
        return None, "Windows 전용"
    creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    try:
        si = subprocess.STARTUPINFO()
        si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    except (AttributeError, TypeError):
        si = None
    try:
        r = subprocess.run(
            ["cmd", "/c", "dir", "/ad", "/b", parent_path],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
            startupinfo=si,
            creationflags=creationflags,
        )
    except subprocess.TimeoutExpired:
        return None, "dir /ad 시간 초과"
    except OSError as e:
        return None, str(e)
    if r.returncode != 0:
        msg = (r.stderr or r.stdout or "").strip() or f"dir 종료 코드 {r.returncode}"
        return None, msg
    names = [ln.strip() for ln in r.stdout.splitlines() if ln.strip()]
    return names, None


def _list_subdir_names_under_parent(parent_path: str):
    """
    제조지침서 루트(parent_path) 바로 아래의 하위 폴더 이름 목록.
    반환: (작업 경로, 이름 목록) 또는 (None, None)
    """
    for candidate in _win_path_access_variants(parent_path):
        try:
            if not os.path.isdir(candidate):
                continue
            out = []
            for n in os.listdir(candidate):
                try:
                    if os.path.isdir(os.path.join(candidate, n)):
                        out.append(n)
                except OSError:
                    continue
            return candidate, out
        except OSError:
            continue
    if os.name == "nt":
        for path_try in _win_path_access_variants(parent_path):
            names, err = _win_dir_list_subdirs_cmd(path_try)
            if names is not None:
                return path_try, names
    return None, None


_MI_CHILD_LEADING_INDEX_RE = re.compile(r"^\d+\.\s*")


def _strip_mi_child_leading_index(name: str) -> str:
    """표준작업지침서 등: '1. BCE01 (96 well)' → 'BCE01 (96 well)'."""
    return _MI_CHILD_LEADING_INDEX_RE.sub("", (name or "").strip(), count=1)


def _child_folder_paren_match(name: str, sub_codes: str) -> bool:
    """BCE01(96) 또는 BCE01 (96 well) 또는 앞에 '1. ' 번호 접두."""
    s = _strip_mi_child_leading_index(name)
    if s.startswith(sub_codes + "("):
        return True
    return re.match(rf"^{re.escape(sub_codes)}\s+\(", s) is not None


def _pick_mi_child_folder_path(parent_work: str, child_names: list, sub_codes: str):
    """
    sub_codes: BCE01 등 UI 코드.
    NAS 실제 폴더명이 BCE01(96-well), BCEPP(PP), '1. BCE01 (96 well)' 인 경우 매칭.
    반환: 자식 폴더 절대 경로 또는 None
    """
    if not child_names:
        return None
    if sub_codes in child_names:
        return os.path.join(parent_work, sub_codes)
    paren = [n for n in child_names if _child_folder_paren_match(n, sub_codes)]
    if len(paren) == 1:
        return os.path.join(parent_work, paren[0])
    if len(paren) > 1:
        paren.sort(key=len)
        return os.path.join(parent_work, paren[0])
    loose = [n for n in child_names if n.startswith(sub_codes)]
    if not loose:
        loose = [
            n
            for n in child_names
            if _strip_mi_child_leading_index(n).startswith(sub_codes)
        ]
    if loose:
        loose.sort(key=len)
        return os.path.join(parent_work, loose[0])
    return None


def _resolve_instruction_subfolder(sub: str, roots: list):
    """
    BCE01 등 하위 폴더에 대해, 주어진 루트 후보 중 첫 접근 가능 경로.
    반환: (folder_path 또는 None, 사용된 루트 또는 None, 시도한 경로 목록)
    """
    tried = []
    for root in roots:
        folder = os.path.normpath(os.path.join(root, sub))
        tried.append(folder)
        for candidate in _win_path_access_variants(folder):
            try:
                if os.path.isdir(candidate):
                    return candidate, root, tried
            except OSError:
                continue
        if os.name == "nt":
            names, _err = _win_dir_list_names_cmd(folder)
            if names is not None:
                return folder, root, tried
            for candidate in list(_win_path_access_variants(folder))[1:]:
                names, _err = _win_dir_list_names_cmd(candidate)
                if names is not None:
                    return candidate, root, tried

        parent_work, subnames = _list_subdir_names_under_parent(root)
        if parent_work and subnames:
            matched = _pick_mi_child_folder_path(parent_work, subnames, sub)
            if matched:
                tried.append(matched)
                for candidate in _win_path_access_variants(matched):
                    try:
                        if os.path.isdir(candidate):
                            return candidate, root, tried
                    except OSError:
                        continue
                if os.name == "nt":
                    n2, _e2 = _win_dir_list_names_cmd(matched)
                    if n2 is not None:
                        return os.path.normpath(matched), root, tried
                    for candidate in list(_win_path_access_variants(matched))[1:]:
                        n2, _e2 = _win_dir_list_names_cmd(candidate)
                        if n2 is not None:
                            return candidate, root, tried
    return None, None, tried


def _resolve_mi_subfolder(sub: str):
    """제조지침서 루트 기준 하위 폴더."""
    return _resolve_instruction_subfolder(sub, _mi_root_candidates())


def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _instruction_nas_hint(env_name: str) -> str:
    return (
        "탐색기에서는 열리는데 여기서만 안 될 때: 서버는 Windows에서 cmd dir로 목록을 한 번 더 시도합니다. "
        "그래도 실패하면 PC를 재로그온한 뒤, 바탕화면의 'cmd'에서 프로젝트 폴더로 이동해 python unified_server.py 로 실행해 보세요 "
        "(Cursor 터미널과 동일 사용자여도 드라이브 핸들이 다를 수 있습니다). "
        f"{env_name}로 탐색기 주소줄 경로를 지정할 수도 있습니다."
    )


def _instruction_latest_payload(roots: list, catalog: tuple, env_hint_name: str) -> dict:
    """제조지침서·표준작업지침서 공통: 모델별 폴더 스캔 결과 JSON dict."""
    out = {
        "base": roots[0] if roots else "",
        "roots_tried": roots,
        "folders": {},
    }
    hint = _instruction_nas_hint(env_hint_name)
    for sub in MANUFACTURING_INSTRUCTION_SUBFOLDERS:
        folder, used_root, tried = _resolve_instruction_subfolder(sub, roots)
        if folder is None:
            out["folders"][sub] = {
                "ok": False,
                "error": "not_found",
                "message": "폴더에 접근할 수 없거나 존재하지 않습니다.",
                "folder_path": tried[-1] if tried else sub,
                "tried_paths": tried,
                "hint": hint,
            }
            continue
        all_docs, _ad_err, _ad_msg = _scan_all_docx_in_folder(folder)
        if all_docs is None:
            out["folders"][sub] = {
                "ok": False,
                "error": _ad_err,
                "message": _ad_msg,
                "folder_path": folder,
                "resolved_base": used_root,
                "tried_paths": tried,
            }
            continue

        best, err_code, err_msg = _scan_latest_docx_by_revision(folder)
        catalog_rows = _build_catalog_rows(sub, all_docs, catalog)
        payload = {
            "ok": True,
            "folder_path": folder,
            "resolved_base": used_root,
            "nas_folder_name": os.path.basename(folder),
            "catalog_rows": catalog_rows,
            "documents": all_docs,
            "documents_count": len(all_docs),
            "tried_paths": tried,
        }
        if best is not None:
            rev, mtime, name, full = best
            payload["revision"] = rev
            payload["filename"] = name
            payload["full_path"] = full
            payload["modified"] = datetime.fromtimestamp(mtime).isoformat(timespec="seconds")
            for d in all_docs:
                d["is_latest_in_folder"] = d["filename"] == name and d.get("full_path") == full
        else:
            payload["revision"] = None
            payload["filename"] = None
            payload["full_path"] = None
            payload["modified"] = None
            payload["latest_note"] = err_msg or "R개정이 있는 최신본 없음"
            for d in all_docs:
                d["is_latest_in_folder"] = False
        out["folders"][sub] = payload
    return out


def _iso_datetime_to_yyyy_mm_dd(s: str) -> str:
    """API/목록용: ISO datetime 문자열 → YYYY-MM-DD."""
    if not s or not isinstance(s, str):
        return ""
    t = s.strip().replace("Z", "+00:00")
    if "T" in t:
        return t.split("T", 1)[0][:10]
    if len(t) >= 10 and t[4] == "-" and t[7] == "-":
        return t[:10]
    return t


def _swi_document_number_prefix_from_filename(filename: str) -> str | None:
    """NAS 원본명 첫 '_' 앞 접두 (예: ESH-DHF-BCE01-SOP-R4)."""
    if not filename or not isinstance(filename, str):
        return None
    base = os.path.splitext(os.path.basename(filename.strip()))[0]
    if not base:
        return None
    if "_" not in base:
        return None
    head = base.split("_", 1)[0].strip()
    return head or None


def _build_swi_catalog_rows(model_code: str, all_docs: list, r_num: int | None) -> list:
    """표준작업지침서: 모델당 1행 (최신 R폴더 기준). 버전은 개정 내용 [버전 x.y] 우선, 없으면 R폴더명."""
    r_folder_label = f"R{r_num}" if r_num is not None else "—"
    fallback_doc_num = f"{model_code} / {r_folder_label}" if r_num is not None else model_code
    if not all_docs:
        return [
            {
                "model_name": model_code,
                "document_number": fallback_doc_num,
                "document_title": "표준 작업 지침서",
                "issue_revision_date": None,
                "version": r_folder_label,
                "matched_filename": None,
            }
        ]
    best = max(all_docs, key=_doc_modified_timestamp)
    fp = best.get("full_path") or ""
    idate = _issued_date_from_docx_headers(fp)
    bracket_ver = _bracket_doc_version_from_revision_docx_path(fp)
    version_display = bracket_ver if bracket_ver else r_folder_label
    prefix = _swi_document_number_prefix_from_filename(best.get("filename") or "")
    doc_num = prefix if prefix else fallback_doc_num
    return [
        {
            "model_name": model_code,
            "document_number": doc_num,
            "document_title": "표준 작업 지침서",
            "issue_revision_date": idate,
            "version": version_display,
            "matched_filename": best["filename"],
        }
    ]


def _instruction_swi_latest_payload(roots: list, env_hint_name: str) -> dict:
    """표준작업지침서: 모델/BCE01 등 → 그 안의 R1,R2,… 중 최대 번호 폴더 → 그 안의 .docx."""
    out = {
        "base": roots[0] if roots else "",
        "roots_tried": roots,
        "folders": {},
        "layout": "swi_model_r_latest_docx",
    }
    hint = _instruction_nas_hint(env_hint_name)
    for sub in MANUFACTURING_INSTRUCTION_SUBFOLDERS:
        folder, used_root, tried = _resolve_instruction_subfolder(sub, roots)
        if folder is None:
            out["folders"][sub] = {
                "ok": False,
                "error": "not_found",
                "message": "폴더에 접근할 수 없거나 존재하지 않습니다.",
                "folder_path": tried[-1] if tried else sub,
                "tried_paths": tried,
                "hint": hint,
            }
            continue

        r_path, r_num = _swi_latest_r_folder_path(folder)
        if not r_path or r_num is None:
            out["folders"][sub] = {
                "ok": False,
                "error": "no_r_folder",
                "message": "모델 폴더 안에 R1, R2, … 형태의 하위 폴더를 찾지 못했습니다.",
                "folder_path": folder,
                "resolved_base": used_root,
                "nas_folder_name": os.path.basename(folder),
                "tried_paths": tried,
                "hint": hint,
            }
            continue

        all_docs, ad_err, ad_msg = _scan_all_docx_in_folder(r_path)
        if all_docs is None:
            out["folders"][sub] = {
                "ok": False,
                "error": ad_err,
                "message": ad_msg,
                "folder_path": folder,
                "swi_latest_r_path": r_path,
                "swi_latest_r_revision": r_num,
                "resolved_base": used_root,
                "tried_paths": tried,
            }
            continue

        for d in all_docs:
            d["swi_r_folder"] = r_num
            d["modified_date"] = _iso_datetime_to_yyyy_mm_dd(d.get("modified") or "")

        catalog_rows = _build_swi_catalog_rows(sub, all_docs, r_num)
        payload = {
            "ok": True,
            "folder_path": folder,
            "swi_latest_r_path": r_path,
            "swi_latest_r_revision": r_num,
            "resolved_base": used_root,
            "nas_folder_name": os.path.basename(folder),
            "catalog_rows": catalog_rows,
            "documents": all_docs,
            "documents_count": len(all_docs),
            "tried_paths": tried,
        }
        if all_docs:
            best = max(all_docs, key=_doc_modified_timestamp)
            payload["filename"] = best["filename"]
            payload["full_path"] = best["full_path"]
            payload["revision"] = r_num
            payload["modified"] = best.get("modified")
            for d in all_docs:
                d["is_latest_in_folder"] = (
                    d["filename"] == best["filename"]
                    and d.get("full_path") == best.get("full_path")
                )
        else:
            payload["revision"] = r_num
            payload["filename"] = None
            payload["full_path"] = None
            payload["modified"] = None
            payload["latest_note"] = "최신 R 폴더에 .docx가 없습니다."
        out["folders"][sub] = payload
    return out


def _qmpc_accessible_base(root: str) -> tuple[str | None, list]:
    """공정도 NAS 루트의 첫 접근 가능한 실제 경로."""
    tried: list = []
    for candidate in _win_path_access_variants(root):
        tried.append(candidate)
        try:
            if os.path.isdir(candidate):
                return os.path.normpath(candidate), tried
        except OSError:
            continue
    if os.name == "nt":
        names, _ = _win_dir_list_names_cmd(root)
        if names is not None:
            return os.path.normpath(root), tried + [root]
    return None, tried


def _qmpc_filename_matches_model(filename: str, model_code: str) -> bool:
    """파일명에 모델(BCE01, BCE PP 등)이 포함되는지 (공정도 루트 평면 구조용)."""
    mc = model_code.replace(" ", "").strip().upper()
    compact = re.sub(r"\s+", "", filename.upper())
    if f"({mc})" in compact:
        return True
    if mc == "BCEPP" and "BCEPP" in compact:
        return True
    # ESH-PC-BCE01은 ESH-PC-BCE010 등과 구분 (BCE01 뒤에 숫자가 이어지면 제외)
    if re.search(rf"ESH-PC-{re.escape(mc)}(?![0-9])", compact):
        return True
    return False


def _qmpc_collect_for_model(model_code: str, roots: list) -> dict:
    """
    모델별 xlsx 후보 수집.
    (1) …/공정도/BCE01/R*/  (구형)
    (2) …/공정도/R*/ 파일명에 모델 포함
    (3) …/공정도/ 바로 아래 파일명에 모델 포함 (R폴더와 동급)
    """
    all_tried: list = []
    hint = _instruction_nas_hint("QMPC_DOC_BASE")
    folder, used_root, t1 = _resolve_instruction_subfolder(model_code, roots)
    all_tried.extend(t1)
    if folder:
        r_path, r_num = _swi_latest_r_folder_path(folder)
        if r_path and r_num is not None:
            docs, _ad_err, _ad_msg = _scan_all_xlsx_in_folder(r_path)
            if docs:
                matched_legacy = [
                    d for d in docs if _qmpc_filename_matches_model(d["filename"], model_code)
                ]
                if matched_legacy:
                    for d in matched_legacy:
                        d["qmpc_r_folder"] = r_num
                        d["modified_date"] = _iso_datetime_to_yyyy_mm_dd(
                            d.get("modified") or ""
                        )
                    return {
                        "ok": True,
                        "documents": matched_legacy,
                        "r_path": r_path,
                        "r_num": r_num,
                        "base_folder": folder,
                        "resolved_base": used_root,
                        "tried_paths": list(all_tried),
                        "layout": "legacy_model_r",
                    }

    for root in roots:
        base, t2 = _qmpc_accessible_base(root)
        all_tried.extend(t2)
        if not base:
            continue
        r_path, r_num = _swi_latest_r_folder_path(base)
        if r_path and r_num is not None:
            docs_r, _e, _m = _scan_all_xlsx_in_folder(r_path)
            if docs_r:
                matched = [d for d in docs_r if _qmpc_filename_matches_model(d["filename"], model_code)]
                if matched:
                    for d in matched:
                        d["qmpc_r_folder"] = r_num
                        d["modified_date"] = _iso_datetime_to_yyyy_mm_dd(d.get("modified") or "")
                    return {
                        "ok": True,
                        "documents": matched,
                        "r_path": r_path,
                        "r_num": r_num,
                        "base_folder": base,
                        "resolved_base": root,
                        "tried_paths": list(all_tried),
                        "layout": "root_r_named",
                    }

        docs0, _e0, _m0 = _scan_all_xlsx_in_folder(base)
        if docs0:
            matched0 = [d for d in docs0 if _qmpc_filename_matches_model(d["filename"], model_code)]
            if matched0:
                r_num_eff = r_num
                r_path_eff = r_path
                if r_num_eff is None:
                    revs = [
                        _max_revision_in_filename(d["filename"]) for d in matched0
                    ]
                    revs2 = [x for x in revs if x is not None]
                    r_num_eff = max(revs2) if revs2 else None
                if r_path_eff is None:
                    r_path_eff = base
                for d in matched0:
                    d["qmpc_r_folder"] = r_num_eff
                    d["modified_date"] = _iso_datetime_to_yyyy_mm_dd(d.get("modified") or "")
                return {
                    "ok": True,
                    "documents": matched0,
                    "r_path": r_path_eff,
                    "r_num": r_num_eff,
                    "base_folder": base,
                    "resolved_base": root,
                    "tried_paths": list(all_tried),
                    "layout": "root_flat_named",
                }

    return {
        "ok": False,
        "error": "not_found",
        "message": "품질관리 공정도 경로에 접근할 수 없거나, 해당 모델의 xlsx를 찾지 못했습니다.",
        "tried_paths": list(all_tried),
        "hint": hint,
    }


def _build_qmpc_catalog_rows(
    model_code: str,
    all_docs: list,
    r_num: int | None,
    qmpc_meta: dict | None = None,
) -> list:
    """품질관리 공정도: 모델당 1행. 문서번호는 엑셀 표지 ESH-PC-… 우선, 없으면 파일명 '_' 앞 접두."""
    r_folder_label = f"R{r_num}" if r_num is not None else "—"
    fallback_doc_num = f"{model_code} / {r_folder_label}" if r_num is not None else model_code
    title = "품질관리 공정도"
    meta = qmpc_meta or {}
    if not all_docs:
        return [
            {
                "model_name": model_code,
                "document_number": fallback_doc_num,
                "document_title": title,
                "issue_revision_date": None,
                "version": r_folder_label,
                "matched_filename": None,
            }
        ]
    best = max(all_docs, key=_doc_modified_timestamp)
    excel_doc = (meta.get("document_number") or "").strip() or None
    prefix = _swi_document_number_prefix_from_filename(best.get("filename") or "")
    doc_num = excel_doc or (prefix if prefix else fallback_doc_num)
    rev_no = meta.get("latest_revision_no")
    if (
        excel_doc
        and rev_no is not None
        and not re.search(r"-R\d+$", excel_doc, re.IGNORECASE)
    ):
        doc_num = f"{excel_doc}-R{rev_no}"
    rev_from_excel = f"R{rev_no}" if rev_no is not None else None
    version_label = rev_from_excel or r_folder_label
    # 제/개정일은 엑셀 개정 이력 표만 사용 (파일 수정 시각 사용 안 함)
    idate = meta.get("latest_revision_date")
    return [
        {
            "model_name": model_code,
            "document_number": doc_num,
            "document_title": title,
            "issue_revision_date": idate,
            "version": version_label,
            "matched_filename": best["filename"],
        }
    ]


def _instruction_qmpc_latest_payload(roots: list, env_hint_name: str) -> dict:
    """품질관리 공정도: BCE01 하위 또는 공정도 루트의 R*/평면 xlsx에서 모델명 매칭."""
    out = {
        "base": roots[0] if roots else "",
        "roots_tried": roots,
        "folders": {},
        "layout": "qmpc_model_r_latest_xlsx",
    }
    hint = _instruction_nas_hint(env_hint_name)
    for sub in MANUFACTURING_INSTRUCTION_SUBFOLDERS:
        col = _qmpc_collect_for_model(sub, roots)
        if not col.get("ok"):
            out["folders"][sub] = {
                "ok": False,
                "error": col.get("error", "not_found"),
                "message": col.get("message", "폴더에 접근할 수 없거나 존재하지 않습니다."),
                "folder_path": (col.get("tried_paths") or [sub])[-1],
                "tried_paths": col.get("tried_paths") or [],
                "hint": col.get("hint") or hint,
            }
            continue

        all_docs = col["documents"]
        r_path = col["r_path"]
        r_num = col["r_num"]
        base_folder = col["base_folder"]
        used_root = col["resolved_base"]
        tried = col.get("tried_paths") or []

        best_for_meta = max(all_docs, key=_doc_modified_timestamp) if all_docs else None
        qmpc_meta = (
            _qmpc_meta_from_xlsx(best_for_meta.get("full_path") or "")
            if best_for_meta
            else {}
        )
        catalog_rows = _build_qmpc_catalog_rows(sub, all_docs, r_num, qmpc_meta=qmpc_meta)
        payload = {
            "ok": True,
            "folder_path": base_folder,
            "qmpc_layout": col.get("layout"),
            "qmpc_latest_r_path": r_path,
            "qmpc_latest_r_revision": r_num,
            "resolved_base": used_root,
            "nas_folder_name": os.path.basename(base_folder),
            "catalog_rows": catalog_rows,
            "documents": all_docs,
            "documents_count": len(all_docs),
            "tried_paths": tried,
        }
        if all_docs and best_for_meta:
            best = best_for_meta
            payload["filename"] = best["filename"]
            payload["full_path"] = best["full_path"]
            payload["revision"] = r_num
            payload["modified"] = best.get("modified")
            for d in all_docs:
                d["is_latest_in_folder"] = (
                    d["filename"] == best["filename"]
                    and d.get("full_path") == best.get("full_path")
                )
        else:
            payload["revision"] = r_num
            payload["filename"] = None
            payload["full_path"] = None
            payload["modified"] = None
            payload["latest_note"] = "매칭된 .xlsx가 없습니다."
        out["folders"][sub] = payload
    return out


# --- 공통 및 정적 파일 서버 ---
@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/styles.css')
def serve_css():
    return send_from_directory('.', 'styles.css')

@app.route('/app.js')
def serve_js():
    return send_from_directory('.', 'app.js')


@app.route('/api/manufacturing_instruction_latest')
def api_manufacturing_instruction_latest():
    """BCE01~BCEPP 각 폴더에서 제조지침서 .docx 조회."""
    out = _instruction_latest_payload(
        _mi_root_candidates(), _MANUFACTURING_INSTRUCTION_CATALOG, "MI_DOC_BASE"
    )
    return jsonify(out)


@app.route("/api/standard_work_instruction_latest")
def api_standard_work_instruction_latest():
    """BCE01~BCEPP: 모델 폴더 내 R* 하위 중 최신 폴더의 표준작업지침서 .docx 1부."""
    out = _instruction_swi_latest_payload(_swi_root_candidates(), "SWI_DOC_BASE")
    return jsonify(out)


@app.route("/api/qmpc_latest")
def api_qmpc_latest():
    """BCE01~BCEPP: 품질관리 공정도 — 모델 폴더 내 R0/R1/… 최신 폴더의 .xlsx."""
    out = _instruction_qmpc_latest_payload(_qmpc_root_candidates(), "QMPC_DOC_BASE")
    return jsonify(out)


def _json_revision_history_response(fk: str, fn: str, path: str | None):
    """개정 이력 표 파싱 결과 Flask 응답."""
    if not path:
        return jsonify(
            {
                "ok": False,
                "error": "not_found",
                "message": "파일을 찾을 수 없거나 허용되지 않는 경로입니다.",
            }
        ), 404
    try:
        with open(path, "rb") as f:
            raw_doc = f.read()
    except OSError as e:
        return jsonify(
            {
                "ok": False,
                "error": "read_error",
                "message": str(e),
            }
        ), 500
    if not raw_doc or len(raw_doc) < 4 or raw_doc[:2] != b"PK":
        return jsonify(
            {
                "ok": False,
                "error": "invalid_docx",
                "message": "docx 형식이 아닙니다.",
            }
        ), 400
    try:
        with zipfile.ZipFile(io.BytesIO(raw_doc), "r") as zf:
            if "word/document.xml" not in zf.namelist():
                rows = []
            else:
                rows = _revision_history_rows_from_document_xml(zf.read("word/document.xml"))
    except (zipfile.BadZipFile, OSError, KeyError, RuntimeError) as e:
        return jsonify(
            {
                "ok": False,
                "error": "parse_error",
                "message": str(e),
            }
        ), 500
    return jsonify(
        {
            "ok": True,
            "folder": fk,
            "filename": fn,
            "rows": rows,
            "row_count": len(rows),
        }
    )


@app.route("/api/manufacturing_instruction_revision_history")
def api_manufacturing_instruction_revision_history():
    """
    제조지침서 .docx 본문의「문서 개정 이력」표 데이터 (여러 표·다음 페이지 이어짐 포함).
    Query: folder=BCE01&filename=...
    """
    fk = (request.args.get("folder") or "").strip()
    fn = (request.args.get("filename") or "").strip()
    if not fk or not fn:
        return jsonify(
            {
                "ok": False,
                "error": "missing_params",
                "message": "folder와 filename 파라미터가 필요합니다.",
            }
        ), 400
    path = _safe_mi_docx_path(fk, fn)
    return _json_revision_history_response(fk, fn, path)


@app.route("/api/standard_work_instruction_revision_history")
def api_standard_work_instruction_revision_history():
    """표준작업지침서 NAS 경로 기준 개정 이력 표. Query: folder=BCE01&filename=..."""
    fk = (request.args.get("folder") or "").strip()
    fn = (request.args.get("filename") or "").strip()
    if not fk or not fn:
        return jsonify(
            {
                "ok": False,
                "error": "missing_params",
                "message": "folder와 filename 파라미터가 필요합니다.",
            }
        ), 400
    path = _safe_swi_docx_path(fk, fn)
    return _json_revision_history_response(fk, fn, path)


@app.route("/api/qmpc_revision_history")
def api_qmpc_revision_history():
    """품질관리 공정도 xlsx 개정 이력 시트(Sheet1 우선) 그리드. Query: folder=BCE01&filename=....xlsx"""
    fk = (request.args.get("folder") or "").strip()
    fn = (request.args.get("filename") or "").strip()
    if not fk or not fn:
        return jsonify(
            {
                "ok": False,
                "error": "missing_params",
                "message": "folder와 filename 파라미터가 필요합니다.",
            }
        ), 400
    path = _safe_qmpc_xlsx_path(fk, fn)
    if not path:
        return jsonify(
            {
                "ok": False,
                "error": "not_found",
                "message": "파일을 찾을 수 없거나 허용되지 않는 경로입니다.",
            }
        ), 404
    grid = _xlsx_second_sheet_revision_grid(path)
    err = grid.get("error")
    if err:
        msg_map = {
            "not_found": "파일을 찾을 수 없습니다.",
            "second_sheet_missing": "개정 이력 시트를 찾지 못했습니다.",
            "revision_sheet_not_found": "개정 이력이 있는 시트를 찾지 못했습니다.",
        }
        msg = msg_map.get(err, "엑셀을 읽을 수 없습니다.")
        if isinstance(err, str) and err.startswith("load_error:"):
            msg = "엑셀 파일을 열 수 없습니다."
        return jsonify({"ok": False, "error": err, "message": msg}), 400
    return jsonify(
        {
            "ok": True,
            "format": "sheet_grid",
            "folder": fk,
            "filename": fn,
            "sheet_name": grid.get("sheet_name") or "",
            "headers": grid.get("headers") or [],
            "rows": grid.get("rows") or [],
            "row_count": len(grid.get("rows") or []),
        }
    )


def _instruction_diag_root_rows(roots: list) -> list:
    rows_out = []
    for root in roots:
        row = {"root": root}
        try:
            row["python_isdir"] = os.path.isdir(root)
        except OSError as e:
            row["python_isdir"] = None
            row["python_isdir_error"] = str(e)
        if os.name == "nt":
            names, err = _win_dir_list_names_cmd(root)
            row["cmd_dir_ok"] = names is not None
            row["cmd_dir_error"] = err if names is None else None
            if names:
                row["cmd_sample_names"] = sorted(names)[:30]
            _pw, subonly = _list_subdir_names_under_parent(root)
            if subonly:
                row["subfolders_bce_match"] = sorted(subonly)[:30]
        rows_out.append(row)
    return rows_out


@app.route("/api/manufacturing_instruction_diag")
def api_manufacturing_instruction_diag():
    """탐색기와 달리 Python이 Z:/UNC를 못 볼 때 원인 확인용 (제조지침서 루트)."""
    import ctypes

    out = {"kind": "manufacturing_instruction", "z_drive_type": None, "z_drive_label": None, "roots": []}
    if os.name == "nt":
        try:
            t = ctypes.windll.kernel32.GetDriveTypeW("Z:\\")
            labels = {0: "UNKNOWN", 1: "NO_ROOT", 2: "REMOVABLE", 3: "FIXED", 4: "REMOTE", 5: "CDROM", 6: "RAMDISK"}
            out["z_drive_type"] = t
            out["z_drive_label"] = labels.get(t, str(t))
        except Exception as e:
            out["z_drive_error"] = str(e)
    out["roots"] = _instruction_diag_root_rows(_mi_root_candidates())
    return jsonify(out)


@app.route("/api/standard_work_instruction_diag")
def api_standard_work_instruction_diag():
    """표준작업지침서 NAS 루트 연결 진단 (JSON)."""
    import ctypes

    out = {"kind": "standard_work_instruction", "z_drive_type": None, "z_drive_label": None, "roots": []}
    if os.name == "nt":
        try:
            t = ctypes.windll.kernel32.GetDriveTypeW("Z:\\")
            labels = {0: "UNKNOWN", 1: "NO_ROOT", 2: "REMOVABLE", 3: "FIXED", 4: "REMOTE", 5: "CDROM", 6: "RAMDISK"}
            out["z_drive_type"] = t
            out["z_drive_label"] = labels.get(t, str(t))
        except Exception as e:
            out["z_drive_error"] = str(e)
    out["roots"] = _instruction_diag_root_rows(_swi_root_candidates())
    return jsonify(out)


@app.route("/api/qmpc_diag")
def api_qmpc_diag():
    """품질관리 공정도 NAS 루트 연결 진단 (JSON)."""
    import ctypes

    out = {"kind": "qmpc", "z_drive_type": None, "z_drive_label": None, "roots": []}
    if os.name == "nt":
        try:
            t = ctypes.windll.kernel32.GetDriveTypeW("Z:\\")
            labels = {0: "UNKNOWN", 1: "NO_ROOT", 2: "REMOVABLE", 3: "FIXED", 4: "REMOTE", 5: "CDROM", 6: "RAMDISK"}
            out["z_drive_type"] = t
            out["z_drive_label"] = labels.get(t, str(t))
        except Exception as e:
            out["z_drive_error"] = str(e)
    out["roots"] = _instruction_diag_root_rows(_qmpc_root_candidates())
    return jsonify(out)


# --- BOM 조회 (Viewer) API ---
@app.route('/api/bom-all')
def get_bom_all():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        tables = ['level0', 'level1', 'level2', 'level3', 'instruction_summary']
        result = {}
        for table in tables:
            if table in ['level1', 'level2', 'level3']:
                name_col = '구성품 명칭' if table == 'level1' else '원재료명'
                cursor.execute(f'''
                    SELECT l.*, i.description as _master_name 
                    FROM {table} l 
                    LEFT JOIN item_master i ON l."코드번호" = i.code_no
                ''')
                rows = []
                for r in cursor.fetchall():
                    d = dict(r)
                    if d.get('_master_name'):
                        d[name_col] = d['_master_name']
                    rows.append(d)
                result[table] = rows
            else:
                cursor.execute(f"SELECT * FROM {table}")
                rows = cursor.fetchall()
                result[table] = [dict(row) for row in rows]
        conn.close()
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/instruction_lots')
def get_instruction_lots():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT DISTINCT "LOT NO.", "제품명", "제조일자" FROM level0 ORDER BY "제조일자" DESC')
        rows = cursor.fetchall()
        conn.close()
        result = [{"lot_no": r["LOT NO."], "product_name": r["제품명"], "mfg_date": r["제조일자"]} for r in rows]
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/instruction_detail/<lot_no>')
def get_instruction_detail(lot_no):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM level0 WHERE "LOT NO." = ?', (lot_no,))
        l0 = cursor.fetchone()
        if not l0:
            conn.close()
            return jsonify({"error": "Lot not found"}), 404
        l0_dict = dict(l0)
        
        cursor.execute('SELECT * FROM level1 WHERE "상위Lot" = ?', (lot_no,))
        l1 = [dict(r) for r in cursor.fetchall()]
        l1_lots = [r["Lot No."] for r in l1 if r["Lot No."]]
        
        l2 = []
        if l1_lots:
            p = ",".join(["?" for _ in l1_lots])
            cursor.execute(f'SELECT * FROM level2 WHERE "상위Lot" IN ({p})', l1_lots)
            l2 = [dict(r) for r in cursor.fetchall()]
        
        l2_lots = [r["Lot No."] for r in l2 if r["Lot No."]]
        l3 = []
        if l2_lots:
            p = ",".join(["?" for _ in l2_lots])
            cursor.execute(f'SELECT * FROM level3 WHERE "상위Lot" IN ({p})', l2_lots)
            l3 = [dict(r) for r in cursor.fetchall()]

        cursor.execute('SELECT * FROM instruction_summary WHERE "상위Lot" = ?', (lot_no,))
        summary = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({"level0": l0_dict, "level1": l1, "level2": l2, "level3": l3, "instruction_summary": summary})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# --- 포장지시서 API ---
@app.route('/api/packaging_preview/<lot_no>')
def get_packaging_preview(lot_no):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM level0 WHERE "LOT NO." = ?', (lot_no,))
        l0 = cursor.fetchone()
        if not l0:
            conn.close()
            return jsonify({"error": "Lot not found"}), 404
        l0 = dict(l0)
        cursor.execute('SELECT * FROM level1 WHERE "상위Lot" = ?', (lot_no,))
        l1_items = [dict(r) for r in cursor.fetchall()]
        cursor.execute('SELECT doc_name FROM instruction_doc_master WHERE code_no = ? AND division LIKE "%PI%"', (l0['제품코드'],))
        doc = cursor.fetchone()
        doc_name = doc['doc_name'] if doc else ""
        cursor.execute(
            'SELECT "포장시 요구량" FROM level1 WHERE "상위Lot" = ? AND UPPER(TRIM("코드번호")) LIKE "EMA015%" LIMIT 1',
            (lot_no,),
        )
        ema_row = cursor.fetchone()
        conn.close()
        try:
            pack_qty = float(str(ema_row[0]).replace(',', '')) if ema_row and ema_row[0] not in (None, '') else None
        except (ValueError, TypeError):
            pack_qty = None
        kit_qty = l0.get('생산 수량(kit)') or 0
        try:
            kit_qty = float(str(kit_qty).replace(',', ''))
        except (ValueError, TypeError):
            kit_qty = 0.0
        total_qty = pack_qty if pack_qty is not None else kit_qty
        return jsonify({
            "E4": doc_name, "A7": l0['제품명'], "J7": l0['제품버전'], "N7": total_qty,
            "S7": l0['제조일자'], "Z7": l0['유효기간'], "AE7": l0['LOT NO.'], "EMA015_items": l1_items
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/packaging_download/<lot_no>')
def download_packaging(lot_no):
    template_path = os.path.join(ROOT_DIR, '25BCE01-포장지시서.xlsx')
    if not os.path.exists(template_path):
        return jsonify({"error": "Template file not found"}), 404
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM level0 WHERE "LOT NO." = ?', (lot_no,))
        l0 = cursor.fetchone()
        if not l0:
            conn.close()
            return jsonify({"error": "Lot not found"}), 404
        l0 = dict(l0)
        cursor.execute('SELECT * FROM level1 WHERE "상위Lot" = ?', (lot_no,))
        l1_items = [dict(r) for r in cursor.fetchall()]
        cursor.execute('SELECT doc_name FROM instruction_doc_master WHERE code_no = ? AND division LIKE "%PI%"', (l0['제품코드'],))
        doc = cursor.fetchone()
        doc_name = doc['doc_name'] if doc else ""
        conn.close()

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        ws['E4']=doc_name; ws['A7']=l0['제품명']; ws['J7']=l0['제품버전']
        ws['S7']=l0['제조일자']; ws['Z7']=l0['유효기간']; ws['AE7']=l0['LOT NO.']
        
        mapping = {'EMA015':21,'EMA014':22,'CR(01)':23,'PC(01)':24,'NC(01)':25,'DA(01)':26,'RD(01)':27,'WS(01)':28,'TM(01)':29,'SS(01)':30,'EMA013':31,'PL(01)':32,'IFU':33}
        try:
            ema015 = next(
                (x for x in l1_items if str(x.get('코드번호') or '').strip().upper().startswith('EMA015')),
                None,
            )
            pq = ema015.get('포장시 요구량') if ema015 else None
            pack_qty = float(str(pq).replace(',', '')) if pq not in (None, '') else None
        except (ValueError, TypeError):
            pack_qty = None
        kq = l0.get('생산 수량(kit)') or 0
        try:
            kq = float(str(kq).replace(',', ''))
        except (ValueError, TypeError):
            kq = 0
        total_qty = pack_qty if pack_qty is not None else kq

        for item in l1_items:
            code = str(item.get('코드번호') or '').strip()
            row_idx = next((row for key, row in mapping.items() if key in code), None)
            try:
                if row_idx and row_idx != 33:
                    ws[f'L{row_idx}'] = item.get('Lot No.')
                    ws[f'S{row_idx}'] = l0['제조일자']
                    ws[f'X{row_idx}'] = item.get('유효기간')
                    ws[f'AI{row_idx}'] = float(str(item.get('포장시 요구량') or 0).replace(',', ''))
            except: pass
            
        ws['L33'] = ''
        ws['S33'] = l0['제조일자']
        ws['X33'] = ''
        ws['AI33'] = total_qty
        ws['N7'] = total_qty

        tmp_fd, tmp_name = tempfile.mkstemp(suffix='.xlsx')
        os.close(tmp_fd)
        wb.save(tmp_name); wb.close()
        return send_file(tmp_name, as_attachment=True, download_name=f'Packaging_Instruction_{lot_no}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# --- 완제품 관리 API ---
@app.route('/api/product_management_preview/<lot_no>')
def get_product_management_preview(lot_no):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM level0 WHERE "LOT NO." = ?', (lot_no,))
        l0 = cursor.fetchone()
        if not l0:
            conn.close()
            return jsonify({"error": "Lot not found"}), 404
        l0 = dict(l0)
        cursor.execute('SELECT "포장시 요구량" FROM level1 WHERE "상위Lot" = ? AND "코드번호" LIKE "EMA015%"', (lot_no,))
        item = cursor.fetchone()
        ema015_qty = item[0] if item else 0
        conn.close()
        return jsonify({
            "A7": l0.get('제품명', ''), "I7": l0.get('제품코드', ''), "N7": l0.get('LOT NO.', ''),
            "T7": l0.get('제조일자', ''), "A9": l0.get('유효기간', ''), "I9": ema015_qty
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/product_management_download/<lot_no>')
def download_product_management(lot_no):
    template_path = os.path.join(ROOT_DIR, '25BCE01-완제품 관리.xlsx')
    if not os.path.exists(template_path):
        return jsonify({"error": "Template file not found"}), 404
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM level0 WHERE "LOT NO." = ?', (lot_no,))
        l0 = cursor.fetchone()
        if not l0:
            conn.close()
            return jsonify({"error": "Lot not found"}), 404
        l0 = dict(l0)
        cursor.execute('SELECT "포장시 요구량" FROM level1 WHERE "상위Lot" = ? AND "코드번호" LIKE "EMA015%"', (lot_no,))
        item = cursor.fetchone()
        ema015_qty = item[0] if item else 0
        conn.close()
        
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        ws['A7']=l0.get('제품명',''); ws['I7']=l0.get('제품코드',''); ws['N7']=l0.get('LOT NO.','')
        ws['T7']=l0.get('제조일자',''); ws['A9']=l0.get('유효기간',''); ws['I9']=ema015_qty
        tmp_fd, tmp_name = tempfile.mkstemp(suffix='.xlsx'); os.close(tmp_fd)
        wb.save(tmp_name); wb.close()
        return send_file(tmp_name, as_attachment=True, download_name=f'Product_Management_{lot_no}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# --- 반제품 관리 API (25BCE01-반제품 관리.xlsx) ---
# B2·M7(item_master unit)·H6~H9·X6·X7, B12:AD30 비움
SEMI_PRODUCT_MGMT_TEMPLATE = '25BCE01-반제품 관리.xlsx'
FINISHED_PRODUCT_MGMT_TEMPLATE = '25BCE01-완제품 관리.xlsx'


def _semi_product_name_line_for_b2(cursor, a7_구성품명칭, code_i7):
    """B2 첫 줄: Level1 구성품 명칭 → 없으면 item_master.description → 없으면 코드."""
    n = str(a7_구성품명칭 or '').strip()
    if n:
        return n
    ku = str(code_i7 or '').strip().upper()
    if not ku:
        return ''
    cursor.execute(
        'SELECT description FROM item_master WHERE UPPER(TRIM(code_no)) = ? LIMIT 1',
        (ku,),
    )
    r = cursor.fetchone()
    if r and r[0] is not None and str(r[0]).strip():
        return str(r[0]).strip()
    return str(code_i7 or '').strip()


# B2 첫 줄: 약어(코드)별 고정 표기 (I7 정규화 키와 일치)
SEMI_B2_DISPLAY_NAME_BY_CODE_KEY = {
    'PB(01)': 'PBSA Buffer',
    'CB(01)': 'Coating Buffer',
    'WB(01)': 'Washing Buffer',
}


def _semi_b2_strip_plate_b_parenthetical(text):
    """문자열에서 (Plate-B) 괄호 구문 제거."""
    s = str(text or '')
    s = re.sub(r'\s*\(\s*Plate-B\s*\)\s*', ' ', s, flags=re.IGNORECASE)
    return re.sub(r'\s{2,}', ' ', s).strip()


def _semi_b2_first_line_display(cursor, a7_구성품명칭, code_i7):
    """B2 표시용 첫 줄: PB/CB/WB 고정명 → PL 계열은 (Plate-B) 제거 → 그 외는 기본 규칙."""
    base = _semi_product_name_line_for_b2(cursor, a7_구성품명칭, code_i7)
    ck = _instruction_code_key(code_i7)
    if ck in SEMI_B2_DISPLAY_NAME_BY_CODE_KEY:
        return SEMI_B2_DISPLAY_NAME_BY_CODE_KEY[ck]
    if ck.startswith('PL'):
        adj = _semi_b2_strip_plate_b_parenthetical(base)
        return adj if adj else base
    return base


def _semi_mgmt_b2_cell_text(cursor, a7_구성품명칭, code_i7):
    """「반제품 명\\n반제품 관리대장」형식."""
    first = _semi_b2_first_line_display(cursor, a7_구성품명칭, code_i7)
    if first:
        return f'{first}\n반제품 관리대장'
    return '반제품 관리대장'


def _item_master_unit(cursor, code_no):
    """item_master.unit (code_no 일치)."""
    ku = str(code_no or '').strip().upper()
    if not ku:
        return ''
    cursor.execute(
        'SELECT unit FROM item_master WHERE UPPER(TRIM(code_no)) = ? LIMIT 1',
        (ku,),
    )
    r = cursor.fetchone()
    if r and r[0] is not None and str(r[0]).strip():
        return str(r[0]).strip()
    return ''


def _semi_mgmt_clear_range_b12_ad30(ws):
    """템플릿 B12:AD30 내용 삭제."""
    for row in range(12, 31):
        for col in range(2, 31):
            _write_cell_safe(ws, f'{get_column_letter(col)}{row}', None)


SEMI_MGMT_FIRST_USAGE_PURPOSE = '성능검사'


def _semi_mgmt_write_usage_history(ws, preview, perf_test_date):
    """
    사용이력 12행부터: B=사용일자, H=사용목적, R=사용량, V=재고량, T·X=단위(동일 값).
    버퍼(PB/CB/WB): level2 사용 이력만(상위 Lot=사용목적, 기존과 동일). 비버퍼: 1행 H=성능검사 + level1 이력 행.
    """
    row = 12
    max_row = 30

    def _write_row(date_v, purpose_v, amt_v, inv_v, unit_v):
        nonlocal row
        if row > max_row:
            return
        _write_cell_safe(ws, f'B{row}', date_v)
        _write_cell_safe(ws, f'H{row}', purpose_v)
        _write_cell_safe(ws, f'R{row}', amt_v)
        _write_cell_safe(ws, f'V{row}', inv_v)
        u = unit_v if unit_v is not None and str(unit_v).strip() != '' else None
        _write_cell_safe(ws, f'T{row}', u)
        _write_cell_safe(ws, f'X{row}', u)
        row += 1

    if preview.get('bufferSemiProduct'):
        unit_fallback = preview.get('M7') or ''
        for item in preview.get('bufferUsageLedger') or []:
            _write_row(
                item.get('usage_date'),
                item.get('usage_purpose'),
                item.get('usage_amount'),
                item.get('inventory_after'),
                item.get('unit') or unit_fallback,
            )
        return

    nb = preview.get('nonBufferLevel1')
    if not nb:
        return
    _write_row(
        (perf_test_date or '').strip() or None,
        SEMI_MGMT_FIRST_USAGE_PURPOSE,
        preview.get('nonBufferPerformanceTestUsage'),
        preview.get('nonBufferInventoryAfterPerfTest'),
        nb.get('unit') or preview.get('M7') or '',
    )
    ufb = nb.get('unit') or preview.get('M7') or ''
    for item in preview.get('nonBufferLevel1LedgerRows') or []:
        _write_row(
            item.get('usage_date'),
            item.get('usage_purpose'),
            item.get('usage_amount'),
            item.get('inventory_after'),
            item.get('unit') or ufb,
        )


def _semi_mgmt_h9_fridge(division_or_code):
    """H9: 약어/코드에 따른 냉장고 자산번호."""
    d = re.sub(r'\s+', '', str(division_or_code or '').strip().upper())
    if d.startswith('PB') or d.startswith('CB') or d.startswith('WB'):
        return '냉장고(ESH-GP-088)'
    for prefix in ('CR', 'PC', 'NC', 'DA', 'PL', 'RD', 'WS', 'TM', 'SS'):
        if d.startswith(prefix):
            return '냉장고(ESH-GP-089)'
    return ''


def _open_semi_mgmt_workbook():
    """반제품 템플릿 우선, 없으면 완제품 관리(동일 입력 셀), 둘 다 없으면 빈 워크북."""
    semi_path = os.path.join(ROOT_DIR, SEMI_PRODUCT_MGMT_TEMPLATE)
    if os.path.exists(semi_path):
        return openpyxl.load_workbook(semi_path)
    fp_path = os.path.join(ROOT_DIR, FINISHED_PRODUCT_MGMT_TEMPLATE)
    if os.path.exists(fp_path):
        return openpyxl.load_workbook(fp_path)
    return openpyxl.Workbook()


def _write_cell_safe(ws, coord, value):
    """병합 셀(MergedCell)이면 병합 범위의 좌상단에만 값을 씁니다."""
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if coord in mr:
                ws.cell(row=mr.min_row, column=mr.min_col, value=value)
                return
        return
    ws[coord] = value


def _split_lot_tokens(s):
    if s is None or str(s).strip() == '':
        return []
    return [x.strip() for x in re.split(r'[\n,;]+', str(s)) if x.strip()]


def _is_buffer_semi_division_or_code(division, code_i7):
    """PB·CB·WB 버퍼류 반제품(약어/코드번호)."""
    for s in (division, code_i7):
        d = re.sub(r'\s+', '', str(s or '').strip().upper())
        if d.startswith('PB') or d.startswith('CB') or d.startswith('WB'):
            return True
    return False


def _level0_production_qty_string(l0):
    """제조기록서(비버퍼)와 동일: 생산 수량(kit) 우선, 없으면 targetQty."""
    if not l0:
        return ''
    kit = l0.get('생산 수량(kit)')
    if kit is not None and str(kit).strip() != '':
        return str(kit).strip()
    alt = l0.get('targetQty')
    if alt is not None and str(alt).strip() != '':
        return str(alt).strip()
    return ''


def _level0_production_qty_float(l0):
    """level0 제조수량 숫자."""
    s = _level0_production_qty_string(l0)
    if not s:
        return 0.0
    try:
        return float(str(s).replace(',', ''))
    except (ValueError, TypeError):
        return 0.0


def _non_buffer_level1_ledger_rows(cursor, parent_lot, lot_tokens, stock_after_perf_test):
    """
    성능검사 반영 후 재고(stock_after_perf_test)에서 시작해,
    동일 상위Lot·반제품 Lot의 level1 행(포장시 요구량)을 순서대로 차감한 사용 이력 행.
    """
    if not parent_lot or not lot_tokens:
        return []

    def _to_float_local(v):
        try:
            return float(str(v or 0).replace(',', ''))
        except (ValueError, TypeError):
            return 0.0

    placeholders = ','.join(['?' for _ in lot_tokens])
    cursor.execute(
        f'''
        SELECT "상위Lot", "제조일자", "유효기간", "구성품 명칭", "코드번호", "포장시 요구량", "단위", "Lot No."
        FROM level1
        WHERE "상위Lot" = ? AND TRIM("Lot No.") IN ({placeholders})
        ORDER BY COALESCE("제조일자", ''), COALESCE("Lot No.", '')
        ''',
        [parent_lot.strip()] + list(lot_tokens),
    )
    out = []
    running = float(stock_after_perf_test) if stock_after_perf_test is not None else 0.0
    pl = (parent_lot or '').strip()
    for r in cursor.fetchall():
        d = dict(r)
        qty = _to_float_local(d.get('포장시 요구량'))
        purpose = str(d.get('상위Lot') or '').strip() or pl
        udate = d.get('제조일자') or d.get('유효기간')
        unit = str(d.get('단위') or '').strip()
        running_after = running - qty
        out.append({
            'usage_date': _fmt_date_yyyy_mm_dd(udate),
            'usage_purpose': purpose,
            'usage_amount': qty,
            'inventory_after': running_after,
            'unit': unit,
        })
        running = running_after
    return out


def _level2_usage_row_sort_key(parent_lot_val):
    """상위Lot 접미(예: …-04R4 → 4)로 사용 순서 추정. 동일 시 문자열로 안정 정렬."""
    s = str(parent_lot_val or '').strip()
    segs = s.split('-')
    if segs:
        last = segs[-1]
        m = re.match(r'^(\d+)', last)
        if m:
            try:
                return (0, int(m.group(1)), s)
            except ValueError:
                pass
    return (1, 0, s)


def _buffer_usage_ledger_from_level2(cursor, lot_tokens, initial_stock, default_unit):
    """
    level2에서 이 버퍼 Lot이 원재료 Lot No.로 할당된 행 = 타 반제품(상위Lot)이 사용한 기록.
    제조량=사용량, 상위Lot=사용 목적(사용처 Lot), 재고량은 제조량 기준 누적 차감.
    """
    if not lot_tokens:
        return []
    def _to_float(v):
        try:
            return float(str(v or 0).replace(',', ''))
        except (ValueError, TypeError):
            return 0.0

    placeholders = ','.join(['?' for _ in lot_tokens])
    cursor.execute(
        f'''
        SELECT "상위Lot", "제조량", "단위", "원재료명", "코드번호", "제조일자", "유효기간"
        FROM level2
        WHERE TRIM("Lot No.") IN ({placeholders})
        AND (
            INSTR(LOWER(COALESCE("원재료명", '')), 'buffer') > 0
            OR UPPER(COALESCE("코드번호", '')) LIKE 'PB%'
            OR UPPER(COALESCE("코드번호", '')) LIKE 'CB%'
            OR UPPER(COALESCE("코드번호", '')) LIKE 'WB%'
        )
        ''',
        lot_tokens,
    )
    rows = sorted(cursor.fetchall(), key=lambda r: (
        str(r['제조일자'] or ''),
        _level2_usage_row_sort_key(r['상위Lot']),
    ))
    out = []
    running = float(initial_stock) if initial_stock is not None else 0.0
    for r in rows:
        d = dict(r)
        used = _to_float(d.get('제조량'))
        unit = str(d.get('단위') or default_unit or '').strip()
        running_after = running - used
        udate = d.get('제조일자') or d.get('유효기간')
        out.append({
            'usage_date': _fmt_date_yyyy_mm_dd(udate),
            'usage_purpose': str(d.get('상위Lot') or '').strip(),
            'usage_amount': used,
            'inventory_after': running_after,
            'unit': unit,
        })
        running = running_after
    return out


def _build_semi_product_management_preview(cursor, parent_lot, semi_lot_raw, division):
    """DB에서 반제품 관리용 필드 조회. 성공 시 (dict, None), 실패 시 (None, error_msg)."""
    parent_lot = (parent_lot or '').strip()
    if not parent_lot:
        return None, 'parent_lot이 필요합니다.'
    semi_lot_raw = semi_lot_raw or ''
    division = (division or '').strip()
    div_u = division.upper()

    cursor.execute('SELECT * FROM level0 WHERE "LOT NO." = ?', (parent_lot,))
    l0r = cursor.fetchone()
    if not l0r:
        return None, '상위 Lot을 찾을 수 없습니다.'
    l0 = dict(l0r)

    cursor.execute('SELECT * FROM level1 WHERE "상위Lot" = ?', (parent_lot,))
    l1_all = [dict(r) for r in cursor.fetchall()]
    cursor.execute('SELECT * FROM instruction_summary WHERE "상위Lot" = ?', (parent_lot,))
    summ_all = [dict(r) for r in cursor.fetchall()]

    semi_tokens = _split_lot_tokens(semi_lot_raw)
    if not semi_tokens and not div_u:
        return None, 'semi_lot 또는 division이 필요합니다.'

    l1 = None

    def lot_matches_row(lot_str):
        if not semi_tokens:
            return False
        lot_str = str(lot_str or '').strip()
        if not lot_str:
            return False
        parts = _split_lot_tokens(lot_str)
        for t in semi_tokens:
            if t == lot_str or t in parts:
                return True
            if t in lot_str or lot_str in t:
                return True
        return False

    for r in l1_all:
        if lot_matches_row(r.get('Lot No.')):
            l1 = r
            break

    if l1 is None and div_u:
        for r in l1_all:
            code = str(r.get('코드번호') or '').strip().upper()
            if not code:
                continue
            if code == div_u or div_u in code or code in div_u:
                l1 = r
                break

    summ = None
    for s in summ_all:
        lot_dot = str(s.get('Lot. No.') or '').strip()
        abbrev = str(s.get('약어') or '').strip().upper()
        if semi_tokens:
            parts = _split_lot_tokens(lot_dot)
            for t in semi_tokens:
                if t == lot_dot or t in parts or (lot_dot and (t in lot_dot or lot_dot in t)):
                    summ = s
                    break
            if summ:
                break
        if div_u and abbrev:
            ak = _instruction_code_key(abbrev)
            dk = _instruction_code_key(div_u)
            if ak == dk or (dk and dk in ak) or (ak and ak in dk):
                summ = s
                break

    if l1 is None and summ is None:
        return None, '반제품에 해당하는 Level1 또는 지시 요약을 찾을 수 없습니다.'

    def _to_float(v):
        try:
            return float(str(v or 0).replace(',', ''))
        except (ValueError, TypeError):
            return 0.0

    i9 = _to_float((l1 or {}).get('포장시 요구량')) if l1 else 0.0
    if i9 == 0 and l1:
        i9 = _to_float(l1.get('할당수량'))

    n7 = str(semi_lot_raw).strip() if str(semi_lot_raw).strip() else ''
    if not n7 and l1:
        n7 = str(l1.get('Lot No.') or '')
    if not n7 and summ:
        n7 = str(summ.get('Lot. No.') or '')

    a7 = str((l1 or {}).get('구성품 명칭') or '')
    i7 = str((l1 or {}).get('코드번호') or '')
    if summ and not a7:
        a7 = str(summ.get('약어') or '')
    if summ and not i7:
        i7 = str(summ.get('약어') or '')
    t7 = ''
    if l1 and l1.get('제조일자'):
        t7 = str(l1.get('제조일자'))
    elif summ and summ.get('제조일자'):
        t7 = str(summ.get('제조일자'))
    else:
        t7 = str(l0.get('제조일자') or '')
    a9 = str((l1 or {}).get('유효기간') or '')

    instr_no = str((summ or {}).get('제조지침서 No.') or '')
    div_out = division
    if summ and summ.get('약어'):
        div_out = str(summ.get('약어'))
    elif not div_out and l1:
        div_out = str(l1.get('코드번호') or '')

    b2 = _semi_mgmt_b2_cell_text(cursor, a7, i7)
    # H6·X6: instruction_summary. H7: 버퍼류는 지시 생산량, 비버퍼는 제조기록서와 동일하게 level0 제조수량
    h6 = x6 = h7 = ''
    if summ:
        h6 = str(summ.get('Lot. No.') or '').strip()
        x6 = _fmt_date_yyyy_mm_dd(summ.get('제조일자'))
        h7 = str(summ.get('생산량') or '').strip()
    if not _is_buffer_semi_division_or_code(div_out, i7):
        l0_qty = _level0_production_qty_string(l0)
        if l0_qty:
            h7 = l0_qty
    mfg_for_x7 = None
    if summ and summ.get('제조일자'):
        mfg_for_x7 = summ.get('제조일자')
    elif l1 and l1.get('제조일자'):
        mfg_for_x7 = l1.get('제조일자')
    else:
        mfg_for_x7 = l0.get('제조일자')
    x7 = _expiry_plus_13_months_minus_1_day(mfg_for_x7)
    if not x7:
        x7 = str(a9 or '').strip()
    h8 = '2 ~ 8℃'
    h9 = _semi_mgmt_h9_fridge(div_out or i7)
    m7 = _item_master_unit(cursor, i7)

    buffer_semi = _is_buffer_semi_division_or_code(div_out, i7)
    buffer_ledger = []
    buffer_ledger_initial = float(i9) if i9 else 0.0
    if buffer_semi and buffer_ledger_initial == 0.0 and h7:
        mqty = re.search(r'[\d,.]+', str(h7))
        if mqty:
            try:
                buffer_ledger_initial = float(mqty.group(0).replace(',', ''))
            except ValueError:
                buffer_ledger_initial = 0.0
    if buffer_semi:
        n7_tokens = _split_lot_tokens(n7) or ([n7.strip()] if str(n7).strip() else [])
        buffer_ledger = _buffer_usage_ledger_from_level2(
            cursor, n7_tokens, buffer_ledger_initial, m7
        )

    nb_l1_info = None
    nb_perf_usage = None
    nb_inv_after_perf = None
    nb_l1_ledger = []
    nb_ledger_initial = None
    if not buffer_semi and l1:
        l0f = _level0_production_qty_float(l0)
        l1pkg = float(i9) if i9 else 0.0
        nb_perf_usage = l0f - l1pkg
        nb_inv_after_perf = l0f - nb_perf_usage
        nb_ledger_initial = l0f
        nb_l1_info = {
            'parent_lot': str(l1.get('상위Lot') or ''),
            'lot_no': str(l1.get('Lot No.') or n7 or ''),
            'code': str(l1.get('코드번호') or i7 or ''),
            'name': str(l1.get('구성품 명칭') or a7 or ''),
            'pack_qty': l1pkg,
            'unit': str((l1.get('단위') or '') or m7 or '').strip(),
            'mfg_date': _fmt_date_yyyy_mm_dd(l1.get('제조일자')),
            'expiry': _fmt_date_yyyy_mm_dd(l1.get('유효기간')),
            'level0_qty': l0f,
        }
        n7_tok_nb = _split_lot_tokens(n7) or ([str(n7).strip()] if str(n7).strip() else [])
        pl_nb = (parent_lot or '').strip()
        if pl_nb and n7_tok_nb and nb_inv_after_perf is not None:
            nb_l1_ledger = _non_buffer_level1_ledger_rows(
                cursor, pl_nb, n7_tok_nb, nb_inv_after_perf
            )

    preview = {
        'A7': a7,
        'I7': i7,
        'N7': n7,
        'T7': t7,
        'A9': a9,
        'I9': i9,
        'B2': b2,
        'M7': m7,
        'H6': h6,
        'X6': x6,
        'H7': h7,
        'X7': x7,
        'H8': h8,
        'H9': h9,
        'division': div_out,
        'instructionNo': instr_no,
        'lotNo': n7,
        'productName': a7,
        'productCode': i7,
        'mfgDate': t7,
        'expiry': x7,
        'qty': i9,
        'bufferSemiProduct': buffer_semi,
        'bufferUsageLedger': buffer_ledger,
        'bufferLedgerInitialStock': buffer_ledger_initial,
        'nonBufferLevel1': nb_l1_info,
        'nonBufferPerformanceTestUsage': nb_perf_usage,
        'nonBufferInventoryAfterPerfTest': nb_inv_after_perf,
        'nonBufferLevel1LedgerRows': nb_l1_ledger,
        'nonBufferLedgerInitialStock': nb_ledger_initial,
    }
    return preview, None


@app.route('/api/semi_product_management_preview')
def semi_product_management_preview():
    parent_lot = request.args.get('parent_lot', '')
    semi_lot = request.args.get('semi_lot', '')
    division = request.args.get('division', '')
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        preview, err = _build_semi_product_management_preview(cursor, parent_lot, semi_lot, division)
        conn.close()
        if err:
            return jsonify({'error': err}), 404
        return jsonify(preview)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/semi_product_management_download')
def semi_product_management_download():
    parent_lot = request.args.get('parent_lot', '')
    semi_lot = request.args.get('semi_lot', '')
    division = request.args.get('division', '')
    include_raw = (request.args.get('include_usage_history') or '').strip().lower()
    include_usage = include_raw in ('1', 'true', 'yes', 'y')
    perf_test_date = (request.args.get('perf_test_date') or '').strip()
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        preview, err = _build_semi_product_management_preview(cursor, parent_lot, semi_lot, division)
        conn.close()
        if err or not preview:
            return jsonify({'error': err or 'No data'}), 404

        wb = _open_semi_mgmt_workbook()
        ws = wb.active
        _semi_mgmt_clear_range_b12_ad30(ws)
        _write_cell_safe(ws, 'B2', preview.get('B2') or '')
        _write_cell_safe(ws, 'M7', preview.get('M7') or '')
        _write_cell_safe(ws, 'H6', preview.get('H6') or '')
        _write_cell_safe(ws, 'X6', preview.get('X6') or '')
        _write_cell_safe(ws, 'H7', preview.get('H7') or '')
        _write_cell_safe(ws, 'X7', preview.get('X7') or '')
        _write_cell_safe(ws, 'H8', preview.get('H8') or '')
        _write_cell_safe(ws, 'H9', preview.get('H9') or '')
        if include_usage:
            _semi_mgmt_write_usage_history(ws, preview, perf_test_date)

        safe_name = re.sub(r'[^\w\-]+', '_', str(preview.get('N7') or 'semi'))[:80]
        suffix = '_usage_history' if include_usage else '_no_usage_history'
        tmp_fd, tmp_name = tempfile.mkstemp(suffix='.xlsx')
        os.close(tmp_fd)
        wb.save(tmp_name)
        wb.close()
        return send_file(
            tmp_name,
            as_attachment=True,
            download_name=f'Semi_Product_Management_{safe_name}{suffix}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# --- 마스터 조회 (제조지시 실행 탭) ---
def _row_get(row, *keys):
    if not row:
        return None
    for k in keys:
        if k in row and row[k] is not None and str(row[k]).strip() != '':
            return row[k]
    return None


@app.route('/api/item_master/<path:code_no>')
def get_item_master(code_no):
    code = urllib.parse.unquote(code_no or '').strip()
    if not code:
        return jsonify({})
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'SELECT code_no, description, detailed_description, version FROM item_master WHERE code_no = ? LIMIT 1',
            (code,),
        )
        r = cursor.fetchone()
        conn.close()
        if not r:
            return jsonify({})
        d = dict(r)
        return jsonify({
            'description': d.get('description') or '',
            'detailed_description': d.get('detailed_description') or '',
            'version': d.get('version') or '',
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/doc_master/<path:code_no>')
def get_doc_master(code_no):
    code = urllib.parse.unquote(code_no or '').strip()
    if not code:
        return jsonify([])
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM instruction_doc_master WHERE code_no = ? ORDER BY id', (code,))
        rows = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify(rows)
    except Exception as e:
        return jsonify([]), 500


def _fmt_date_yyyy_mm_dd(val):
    """제조일자를 YYYY-MM-DD 문자열로 통일 (저장 시). YYMMDD(6자리) → 20YY-MM-DD."""
    if val is None or val == '':
        return ''
    s = str(val).strip()
    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
        return s[:10]
    digits = re.sub(r'\D', '', s)
    if len(digits) == 6:
        return f'20{digits[:2]}-{digits[2:4]}-{digits[4:6]}'
    if len(digits) >= 8:
        return f'{digits[:4]}-{digits[4:6]}-{digits[6:8]}'
    return s[:10] if len(s) >= 10 else s


def _parse_mfg_date_to_date(val):
    """app.js parseDateInput 대응: 숫자 8자리(YYYYMMDD), 6자리(YYMMDD), YYYY-MM-DD."""
    if val is None or str(val).strip() == '':
        return None
    s = str(val).strip()
    digits = re.sub(r'\D', '', s)
    if len(digits) >= 8:
        try:
            return date(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))
        except ValueError:
            return None
    if len(digits) == 6:
        try:
            return date(2000 + int(digits[:2]), int(digits[2:4]), int(digits[4:6]))
        except ValueError:
            return None
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def _expiry_plus_13_months_minus_1_day(mfg_val):
    """
    반제품 유효기간: 제조일 + 13개월(말일 보정) − 1일.
    app.js calcExpiryDate(setMonth+13), setDate(0) 보정, setDate(-1) 와 동치.
    """
    mfg = _parse_mfg_date_to_date(mfg_val)
    if not mfg:
        return ''
    y, mo, d = mfg.year, mfg.month, mfg.day
    total_m = y * 12 + (mo - 1) + 13
    ny = total_m // 12
    nm = total_m % 12 + 1
    last = calendar.monthrange(ny, nm)[1]
    nd = min(d, last)
    d2 = date(ny, nm, nd)
    d3 = d2 - timedelta(days=1)
    return d3.strftime('%Y-%m-%d')


def _lot_no_equiv_set(lot_str):
    """
    Lot 문자열 동치(날짜 접두만 다른 경우): 111127-01PB-01R3 ↔ 2011-11-27-01PB-01R3
    """
    s = str(lot_str or '').strip()
    if not s:
        return frozenset()
    out = {s}
    m = re.match(r'^(\d{2})(\d{2})(\d{2})(-.+)$', s)
    if m:
        yy, mm, dd, rest = m.groups()
        out.add(f'20{yy}-{mm}-{dd}{rest}')
    m = re.match(r'^(20\d{2})-(\d{2})-(\d{2})(-.+)$', s)
    if m:
        y, mm, dd, rest = m.group(1), m.group(2), m.group(3), m.group(4)
        out.add(f'{y[2:]}{mm}{dd}{rest}')
    return frozenset(out)


def _lot_refs_equal(a, b):
    """instruction_summary Lot. No. 와 level3 상위Lot이 같은 반제품을 가리키는지."""
    a, b = str(a or '').strip(), str(b or '').strip()
    if not a or not b:
        return False
    if a == b:
        return True
    return bool(_lot_no_equiv_set(a) & _lot_no_equiv_set(b))


def _gubun_from_item_master(cursor, code_no, memo):
    """item_master의 category로 구분(완제품·반제품·원재료·소모품 등) 조회. IFU 계열은 소모품."""
    k = str(code_no or '').strip()
    if not k:
        return ''
    ku = k.upper()
    if ku == 'IFU' or ku.startswith('IFU'):
        return '소모품'
    if ku in memo:
        return memo[ku]
    cursor.execute(
        'SELECT category FROM item_master WHERE UPPER(TRIM(code_no)) = ? LIMIT 1',
        (ku,),
    )
    row = cursor.fetchone()
    cat = ''
    if row and row[0] is not None:
        cat = str(row[0]).strip()
    memo[ku] = cat
    return cat


def _manufacturer_from_item_master(cursor, code_no, memo):
    """item_master의 manufacturer를 code_no로 조회."""
    k = str(code_no or '').strip()
    if not k:
        return ''
    ku = k.upper()
    if ku in memo:
        return memo[ku]
    cursor.execute(
        'SELECT manufacturer FROM item_master WHERE UPPER(TRIM(code_no)) = ? LIMIT 1',
        (ku,),
    )
    row = cursor.fetchone()
    mfr = ''
    if row and row[0] is not None:
        mfr = str(row[0]).strip()
    memo[ku] = mfr
    return mfr


def _instruction_code_key(s):
    """약어·코드번호 비교용(공백 제거, 대문자)."""
    return re.sub(r'\s+', '', str(s or '').strip().upper())


def _l1_row_for_instruction_summary(l1_raw, parent_lot, division, semi_lot):
    """
    Level1에서 상위Lot=생산Lot, 코드번호=약어(division),
    할당 Lot이 반제품 Lot(calcLot)과 맞는 첫 행(없으면 None).
    """
    parent_lot = (parent_lot or '').strip()
    div_key = _instruction_code_key(division)
    semi_s = str(semi_lot or '').strip()
    semi_tokens = _split_lot_tokens(semi_lot)
    if not parent_lot or not div_key or not semi_s:
        return None
    for r in l1_raw:
        pl = str(_row_get(r, '상위Lot', '상위 Lot', '상위 LOT') or '').strip()
        if pl != parent_lot:
            continue
        code = str(_row_get(r, '코드번호', 'Code No.', 'Code') or '').strip().upper()
        if _instruction_code_key(code) != div_key:
            continue
        lot_cell = str(_row_get(r, 'Lot No.', '할당 Lot', '할당Lot') or '').strip()
        lot_set = set(_split_lot_tokens(lot_cell))
        if lot_cell:
            lot_set.add(lot_cell)
        ok = False
        for t in semi_tokens:
            if not t:
                continue
            if t == lot_cell or t in lot_set:
                ok = True
                break
            if lot_cell and (t in lot_cell or lot_cell in t):
                ok = True
                break
        if not ok:
            continue
        return r
    return None


def _l1_packaging_qty_for_instruction_summary(l1_raw, parent_lot, division, semi_lot):
    """
    instruction_summary 생산량: Level1에서 상위Lot=생산Lot, 코드번호=약어(division),
    할당 Lot이 반제품 Lot(calcLot)과 맞는 행의 포장시 요구량(또는 할당수량 등).
    """
    r = _l1_row_for_instruction_summary(l1_raw, parent_lot, division, semi_lot)
    if not r:
        return ''
    q = _row_get(r, '포장시 요구량', '할당수량', '필요 수량', '제조량')
    if q is None or str(q).strip() == '':
        return ''
    return str(q).strip()


def _l3_cam006_alloc_for_instruction_lot(l3_raw, instruction_lot):
    """
    업로드 CSV level3: PB/CB/WB 반제품 Lot(calcLot)과 상위Lot이 동치(날짜 접두 포함)인 행 중
    코드번호 CAM006(또는 자료에 CMA006으로 적힌 경우)의 할당수량만 사용(첫 건, 합산 없음).
    """
    semi = str(instruction_lot or '').strip()
    if not semi:
        return ''
    cam_keys = frozenset({'CAM006', 'CMA006'})
    for r in l3_raw:
        code = str(_row_get(r, '코드번호', 'Code No.', 'Code') or '').strip().upper()
        if _instruction_code_key(code) not in cam_keys:
            continue
        pl3 = str(_row_get(r, '상위Lot', '상위 Lot', '상위 LOT') or '').strip()
        if not _lot_refs_equal(semi, pl3):
            continue
        q = _row_get(r, '할당수량')
        if q is not None and str(q).strip() != '':
            return str(q).strip()
    return ''


def _l1_packaging_qty_for_cr(l1_raw, parent_lot):
    """
    Level1에서 상위Lot=생산 Lot이고 코드번호가 CR로 시작하는 첫 행의 포장시 요구량.
    PI instruction_summary 생산량을 CR과 동일하게 맞출 때 사용(calcLot 없음).
    """
    parent_lot = (parent_lot or '').strip()
    if not parent_lot:
        return ''
    for r in l1_raw:
        pl = str(_row_get(r, '상위Lot', '상위 Lot', '상위 LOT') or '').strip()
        if pl != parent_lot:
            continue
        code = str(_row_get(r, '코드번호', 'Code No.', 'Code') or '').strip().upper()
        if not code.startswith('CR'):
            continue
        q = _row_get(r, '포장시 요구량', '할당수량', '필요 수량', '제조량')
        if q is None or str(q).strip() == '':
            continue
        return str(q).strip()
    return ''


@app.route('/api/save_instruction', methods=['POST'])
def save_instruction():
    try:
        data = request.get_json(force=True, silent=True) or {}
        l0_src = data.get('level0') or {}
        lot_no = (l0_src.get('lotNo') or l0_src.get('LOT NO.') or '').strip()
        if not lot_no:
            return jsonify({'status': 'error', 'error': 'LOT No.가 없습니다.'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()
        cat_memo = {}
        mfr_memo = {}

        def gubun_for_code(code_no, fallback=''):
            g = _gubun_from_item_master(cursor, code_no, cat_memo)
            return g if g else (fallback or '')

        def manufacturer_for_code(code_no, fallback=''):
            m = _manufacturer_from_item_master(cursor, code_no, mfr_memo)
            return m if m else (fallback or '')

        l0_code = l0_src.get('modelName') or l0_src.get('제품코드') or ''
        l0 = {
            'Level': 0,
            '제품코드': l0_code,
            '구분': gubun_for_code(l0_code, l0_src.get('구분') or ''),
            '제품명': l0_src.get('productName') or l0_src.get('제품명') or '',
            'LOT NO.': lot_no,
            '제품버전': l0_src.get('version') or l0_src.get('제품버전') or '',
            '제조일자': _fmt_date_yyyy_mm_dd(l0_src.get('mfgDate') or l0_src.get('제조일자') or ''),
            '유효기간': l0_src.get('expiryDate') or l0_src.get('유효기간') or '',
            '생산 수량(kit)': l0_src.get('targetQty') or l0_src.get('생산 수량(kit)') or '',
            '생산의뢰서 번호': l0_src.get('생산의뢰서 번호') or '',
            '의뢰팀': l0_src.get('requestTeam') or l0_src.get('의뢰팀') or '',
            '생산목적': l0_src.get('purpose') or l0_src.get('생산목적') or '',
            '작업자': l0_src.get('작업자') or '',
            '검사자': l0_src.get('검사자') or '',
            '검사일': l0_src.get('검사일') or '',
            '완제품검사 문서번호': l0_src.get('완제품검사 문서번호') or '',
            '제품정보': l0_src.get('productInfo') or l0_src.get('제품정보') or '',
        }

        def row_mfg_date(row):
            raw = _row_get(row, '제조일자', '제조 일자', 'MfgDate', 'mfgDate', 'Manufacturing Date') or ''
            fmt = _fmt_date_yyyy_mm_dd(raw)
            return fmt if fmt else l0['제조일자']

        def norm_l1(row):
            code = _row_get(row, '코드번호', 'Code No.', 'Code') or ''
            return {
                'Level': int(_row_get(row, 'Level', 'level') or 1),
                '상위Lot': _row_get(row, '상위Lot', '상위 Lot', '상위 LOT') or '',
                '코드번호': code,
                '구분': gubun_for_code(code, _row_get(row, '구분') or ''),
                '구성품 명칭': _row_get(row, '구성품 명칭', '명칭 / 구성품', '명칭/구성품') or '',
                'Lot No.': _row_get(row, 'Lot No.', '할당 Lot', '할당Lot') or '',
                '제조일자': row_mfg_date(row),
                '유효기간': _row_get(row, '유효기간') or '',
                '포장 기준량': _row_get(row, '포장 기준량') or '',
                '포장시 요구량': _row_get(row, '포장시 요구량', '할당수량', '필요 수량', '제조량') or '',
                '단위': _row_get(row, '단위') or '',
            }

        def norm_l2(row):
            code = _row_get(row, '코드번호', 'Code No.', 'Code') or ''
            return {
                'Level': int(_row_get(row, 'Level', 'level') or 2),
                '상위Lot': _row_get(row, '상위Lot', '상위 Lot', '상위 LOT') or '',
                '코드번호': code,
                '구분': gubun_for_code(code, _row_get(row, '구분') or ''),
                '원재료명': _row_get(row, '원재료명', '명칭 / 구성품', '구성품 명칭') or '',
                '제조사': manufacturer_for_code(code, _row_get(row, '제조사', 'Maker', 'maker', 'Manufacturer') or ''),
                'Lot No.': _row_get(row, 'Lot No.', '할당 Lot', '할당Lot') or '',
                '제조일자': row_mfg_date(row),
                '유효기간': _row_get(row, '유효기간') or '',
                '제조량': _row_get(row, '제조량', '할당수량', '포장시 요구량', '필요 수량') or '',
                '단위': _row_get(row, '단위') or '',
            }

        l1_raw = data.get('level1') or []
        l2_raw = data.get('level2') or []
        l3_raw = data.get('level3') or []
        l1_rows = [norm_l1(r) for r in l1_raw]
        l2_rows = [norm_l2(r) for r in l2_raw]
        l3_rows = [norm_l2(r) for r in l3_raw]

        summary_in = data.get('instruction_summary') or []
        summary_rows = []
        for item in summary_in:
            div = item.get('division') or item.get('약어') or ''
            calc_raw = item.get('calcLot') or item.get('Lot. No.') or ''
            div_u = (div or '').strip().upper()
            is_pb_cb_wb = (
                div_u.startswith('PB') or div_u.startswith('CB') or div_u.startswith('WB')
            )
            if is_pb_cb_wb:
                prod_qty = _l3_cam006_alloc_for_instruction_lot(l3_raw, calc_raw)
            else:
                prod_qty = ''
                qty_l1 = _l1_packaging_qty_for_instruction_summary(l1_raw, lot_no, div, calc_raw)
                if not qty_l1 and div_u.startswith('PI'):
                    qty_l1 = _l1_packaging_qty_for_cr(l1_raw, lot_no)
                prod_qty = qty_l1 if qty_l1 else (item.get('생산량') or item.get('productionQty') or '')
            summary_rows.append({
                '상위Lot': lot_no,
                '약어': div,
                '제조지침서 No.': item.get('latest_doc_no') or item.get('제조지침서 No.') or '',
                'Lot. No.': str(calc_raw).strip(),
                '생산량': prod_qty,
                '제조일자': _fmt_date_yyyy_mm_dd(
                    item.get('mfgDate') or item.get('제조일자') or l0_src.get('mfgDate') or l0_src.get('제조일자')
                ),
            })

        def insert_table(cursor, table, rows):
            if not rows:
                return
            cols = list(rows[0].keys())
            qcols = ','.join(['"' + c.replace('"', '""') + '"' for c in cols])
            ph = ','.join(['?' for _ in cols])
            sql = f'INSERT INTO {table} ({qcols}) VALUES ({ph})'
            for r in rows:
                cursor.execute(sql, [r[c] for c in cols])

        cursor.execute(
            'DELETE FROM level3 WHERE "상위Lot" IN (SELECT "Lot No." FROM level2 WHERE "상위Lot" IN (SELECT "Lot No." FROM level1 WHERE "상위Lot" = ?))',
            (lot_no,),
        )
        cursor.execute(
            'DELETE FROM level2 WHERE "상위Lot" IN (SELECT "Lot No." FROM level1 WHERE "상위Lot" = ?)',
            (lot_no,),
        )
        cursor.execute('DELETE FROM level1 WHERE "상위Lot" = ?', (lot_no,))
        cursor.execute('DELETE FROM instruction_summary WHERE "상위Lot" = ?', (lot_no,))
        cursor.execute('DELETE FROM level0 WHERE "LOT NO." = ?', (lot_no,))

        insert_table(cursor, 'level0', [l0])
        insert_table(cursor, 'level1', l1_rows)
        insert_table(cursor, 'level2', l2_rows)
        insert_table(cursor, 'level3', l3_rows)
        insert_table(cursor, 'instruction_summary', summary_rows)

        conn.commit()
        conn.close()
        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500


if __name__ == '__main__':
    print("--- Unified BOM System Server Starting ---")
    # Windows: debug 리로더 자식 프로세스에서 Z: 등 네트워크 드라이브가 사라지는 경우가 있어 기본 비활성화.
    _reload = os.environ.get("BOM_USE_RELOADER", "").strip().lower() in ("1", "true", "yes", "y")
    app.run(host="0.0.0.0", port=9000, debug=True, use_reloader=_reload)
